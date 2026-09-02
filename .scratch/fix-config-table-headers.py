# -*- coding: utf-8 -*-
"""Fix malformed Excel three-row headers (row1 English duplicate) and rebake CSV."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / ".scratch" / "tools"))
from bake_config_tables_py import build_csv, excel_to_csv_base, find_header_index, is_english_header, read_sheet_rows

EXCEL_DIR = ROOT / "Gravedigger2026" / "Assets" / "ConfigTables" / "Excel"
CSV_DIR = ROOT / "Gravedigger2026" / "Assets" / "ConfigTables" / "Csv"

TABLE_HEADERS: dict[str, tuple[list[str], list[str]]] = {
    "Manufacture_MagicBookConfig": (
        [
            "魔法书ID",
            "是否唯一",
            "概率型",
            "生效环节",
            "魔法书效果",
            "魔法书效果参数",
            "魔法书Icon",
            "魔法书名称",
            "魔法书介绍",
            "特效外观ID",
            "特效优先级",
            "特效强度加算",
        ],
        [
            "主键",
            "1=同 Id 不可叠装第二本；0=默认可叠（各占一槽）",
            "1=概率触发魔法；0=无概率；空=0",
            "触发时机；可多值",
            "登记制 PascalCase Token；空=无效果",
            "与 Token 配套的参数；空=无参/缺省",
            "UI 图标资源 Id",
            "展示名；若启用 i18n 可为 Key",
            "展示文案",
            "空=该书无特效外观；非空 → Catalog StyleId",
            "缺/空=0；仅材质通道",
            "缺/空=1；材质通道或放大通道",
        ],
    ),
    "Protagonist_ProtagonistEquipmentConfig": (
        [
            "装备ID",
            "装备等级",
            "装备名称",
            "装备图标",
            "升下一级经验",
            "转化经验值",
            "装备生效功能",
            "装备效果",
            "装备描述",
        ],
        [
            "复合主键之一",
            "复合主键之一；从 1 起",
            "展示名；若启用 i18n 可为 Key",
            "UI 图标资源 Id",
            "升到 EquipLevel+1 所需；空或 ≤0 → 该行为满级",
            "再获同 EquipId 时转入的经验",
            "Dig | SoldierManufacture | Combat；可多值",
            "Dig 域与科技 AttributeModifiers 同风格；空=无效果",
            "展示文案",
        ],
    ),
    "Combat_MonsterSkillEffectConfig": (
        ["怪物技能ID", "名称", "效果种类", "效果参数"],
        [
            "主键",
            "展示名",
            "登记制；首版 MonsterSelfReviveOnDeath",
            "DelaySeconds / ReviveHpRatio / InvincibleSeconds 等",
        ],
    ),
    "Item_ItemCatalogConfig": (
        ["道具ID", "道具名", "道具图标", "道具类型", "管理道具属性配置表", "道具描述", "售卖价格"],
        [
            "主键；奖励系统公共 Id",
            "道具公共展示名；空则 UI 回退 ItemId",
            "奖励弹窗/通用掉落 UI 图标资源 Id",
            "Currency | Material | BodyPart | MagicBook | ProtagonistEquipment",
            "该道具权威来源表名；运行时据此校验与分发",
            "通用描述文案",
            "≥ 0；商店出售入账精魂",
        ],
    ),
    "Audio_BgmConfig": (
        ["BGM行ID", "情境", "音频资源ID", "是否循环", "随机权重", "音量"],
        [
            "主键",
            "Title | Dig | Combat",
            "与 BgmClipCatalog 键一致；源文件 Art/Audio/BGM/",
            "缺省/空=1（循环）；0=只播一遍",
            "同 Context 加权；0 剔除；缺省/空=1",
            "0～1；缺省/空=1",
        ],
    ),
}


def find_english_header_row(ws) -> tuple[int, list[str]]:
    for row_idx in range(1, 4):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        while values and (values[-1] is None or str(values[-1]).strip() == ""):
            values.pop()
        text = [str(v).strip() if v is not None else "" for v in values]
        if is_english_header(text):
            return row_idx, text
    raise ValueError("no English header in first 3 rows")


def fix_table(xlsx_path: Path) -> None:
    csv_base = excel_to_csv_base(xlsx_path.stem)
    if csv_base not in TABLE_HEADERS:
        raise KeyError(f"no metadata for {csv_base}")

    row1_zh, row2_desc = TABLE_HEADERS[csv_base]

    with open(xlsx_path, "rb") as f:
        data = f.read()

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    header_row_idx, english_cols = find_english_header_row(ws)

    if header_row_idx == 1 and ws.cell(2, 1).value and "待补 SPEC" in str(ws.cell(2, 1).value):
        # Broken hybrid: row1 English duplicate + row2 placeholder + row3 English.
        if ws.cell(3, 1).value == english_cols[0]:
            header_row_idx = 3

    if len(row1_zh) != len(english_cols) or len(row2_desc) != len(english_cols):
        raise ValueError(
            f"{xlsx_path.name}: metadata column count mismatch "
            f"(excel={len(english_cols)}, zh={len(row1_zh)}, desc={len(row2_desc)})"
        )

    for col, val in enumerate(row1_zh, 1):
        ws.cell(1, col, val)
    for col, val in enumerate(row2_desc, 1):
        ws.cell(2, col, val)
    for col, val in enumerate(english_cols, 1):
        ws.cell(3, col, val)

    wb.save(xlsx_path)
    wb.close()

    rows = read_sheet_rows(xlsx_path)
    header_i = find_header_index(rows)
    if header_i != 2:
        raise ValueError(f"{xlsx_path.name}: expected header at row 3 (index 2), got {header_i}")

    csv_text = build_csv(rows[header_i:])
    out = CSV_DIR / f"{csv_base}.csv"
    out.write_text(csv_text, encoding="utf-8", newline="\n")
    line_count = len([line for line in csv_text.splitlines() if line.strip()])
    print(f"FIXED {xlsx_path.name} -> {out.name} ({line_count} lines, header_index={header_i})")


def main() -> None:
    for xlsx in sorted(EXCEL_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        csv_base = excel_to_csv_base(xlsx.stem)
        if csv_base not in TABLE_HEADERS:
            continue
        fix_table(xlsx)


if __name__ == "__main__":
    main()
