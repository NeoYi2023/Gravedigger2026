import csv
import io
import re
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1] / "Gravedigger2026" / "Assets" / "ConfigTables"
TARGETS = [
    (
        ROOT / "Excel" / "通用_常量表_Combat_CombatConstantConfig.xlsx",
        ROOT / "Csv" / "Combat_CombatConstantConfig.csv",
    ),
    (
        ROOT / "Mode2" / "Excel" / "通用_常量表_Combat_CombatConstantConfig.xlsx",
        ROOT / "Mode2" / "Csv" / "Combat_CombatConstantConfig.csv",
    ),
]
EXPECTED_R1 = ["常量键", "主键中文翻译", "数值", "备注", "备注中文解释"]
EXPECTED_R2 = ["主键", "展示用中文名；运行时不读", "float", "英文备注；运行时不读", "中文说明；运行时不读"]
EXPECTED_R3 = ["ConstantKey", "ConstantKeyZh", "Value", "Comment", "CommentZh"]
KNOCKBACK_KEYS = {
    "DeathKnockbackDirectionSpreadHalfDegrees",
    "DeathKnockbackDirectionRandomStepDegrees",
}
EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")

def read_xlsx_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    wb.close()
    return rows


def is_english_header(row):
    saw = False
    for cell in row:
        if not cell.strip():
            continue
        saw = True
        if not EN_COL.match(cell.strip()):
            return False
    return saw


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in (",", '"', "\n", "\r")):
        return '"' + value.replace('"', '""') + '"'
    return value


def bake_to_csv_text(rows):
    header_idx = next(i for i in range(min(3, len(rows))) if is_english_header(rows[i]))
    lines = []
    for row in rows[header_idx:]:
        if not any(c.strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c) for c in row))
    return "\n".join(lines) + "\n"


def main():
    for xlsx_path, csv_path in TARGETS:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        assert [ws.cell(1, c).value for c in range(1, 6)] == EXPECTED_R1
        assert [ws.cell(2, c).value for c in range(1, 6)] == EXPECTED_R2
        assert [ws.cell(3, c).value for c in range(1, 6)] == EXPECTED_R3
        found = set()
        for r in range(4, ws.max_row + 1):
            key = ws.cell(r, 1).value
            if key in KNOCKBACK_KEYS:
                found.add(key)
        assert found == KNOCKBACK_KEYS
        wb.close()
        print("header ok:", xlsx_path.name)

        baked = list(csv.reader(io.StringIO(bake_to_csv_text(read_xlsx_rows(xlsx_path)))))
        current = list(csv.reader(csv_path.open("r", encoding="utf-8-sig", newline="")))
        assert baked == current, f"bake mismatch for {csv_path.name}"
        print("bake semantic match:", csv_path.name)

    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
