# -*- coding: utf-8 -*-
"""TF-06: fill Form_Wedge_01 StatModifiers and bake Mode1+Mode2 CSV (SPEC_04 §14.7)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tf01_tactical_formation_config import (
    FORMATION_CSV,
    FORMATION_XLSX,
    ROOT,
    bake,
    header_map,
)

STAT_MODIFIERS = "Stat=Strength|Mul=1.15"
FORMATION_ID = "Form_Wedge_01"


def patch_stat(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    cols = header_map(ws)
    id_col = cols["FormationId"]
    stat_col = cols["StatModifiers"]
    found = False
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, id_col).value
        if sid is not None and str(sid).strip() == FORMATION_ID:
            ws.cell(r, stat_col, STAT_MODIFIERS)
            found = True
            print(f"  {xlsx.name} row {r} StatModifiers={STAT_MODIFIERS}")
            break
    if not found:
        raise SystemExit(f"{xlsx}: missing {FORMATION_ID}")
    wb.save(xlsx)


def main() -> None:
    pairs = [
        (ROOT / "Excel", ROOT / "Csv"),
        (ROOT / "Mode2" / "Excel", ROOT / "Mode2" / "Csv"),
    ]
    for excel_dir, csv_dir in pairs:
        xlsx = excel_dir / FORMATION_XLSX
        patch_stat(xlsx)
        bake(xlsx, csv_dir / FORMATION_CSV)


if __name__ == "__main__":
    main()
