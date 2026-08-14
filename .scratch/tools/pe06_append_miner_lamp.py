# -*- coding: utf-8 -*-
"""PE-06: append Equip_MinerLamp L1-5 to ProtagonistEquipmentConfig Excel (Mode1+Mode2)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
XLSX_NAME = "主角_装备配置表_Protagonist_ProtagonistEquipmentConfig.xlsx"

LAMP_ROWS = [
    [
        "Equip_MinerLamp",
        "1",
        "矿灯",
        "Icon_MinerLamp",
        "1",
        "1",
        "Dig",
        "GraveSpawnWeightBonus_Q4_10|GraveSpawnWeightBonus_Q5_10|GraveSpawnWeightBonus_Q6_10",
        "每级增加Q4/Q5/Q6生成权重10（1级）",
    ],
    [
        "Equip_MinerLamp",
        "2",
        "矿灯",
        "Icon_MinerLamp",
        "1",
        "1",
        "Dig",
        "GraveSpawnWeightBonus_Q4_20|GraveSpawnWeightBonus_Q5_20|GraveSpawnWeightBonus_Q6_20",
        "每级增加Q4/Q5/Q6生成权重10（2级）",
    ],
    [
        "Equip_MinerLamp",
        "3",
        "矿灯",
        "Icon_MinerLamp",
        "1",
        "1",
        "Dig",
        "GraveSpawnWeightBonus_Q4_30|GraveSpawnWeightBonus_Q5_30|GraveSpawnWeightBonus_Q6_30",
        "每级增加Q4/Q5/Q6生成权重10（3级）",
    ],
    [
        "Equip_MinerLamp",
        "4",
        "矿灯",
        "Icon_MinerLamp",
        "1",
        "1",
        "Dig",
        "GraveSpawnWeightBonus_Q4_40|GraveSpawnWeightBonus_Q5_40|GraveSpawnWeightBonus_Q6_40",
        "每级增加Q4/Q5/Q6生成权重10（4级）",
    ],
    [
        "Equip_MinerLamp",
        "5",
        "矿灯",
        "Icon_MinerLamp",
        "",
        "1",
        "Dig",
        "GraveSpawnWeightBonus_Q4_50|GraveSpawnWeightBonus_Q5_50|GraveSpawnWeightBonus_Q6_50",
        "每级增加Q4/Q5/Q6生成权重10（满级）",
    ],
]


def already_has_lamp(ws) -> bool:
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        if row and row[0] == "Equip_MinerLamp":
            return True
    return False


def append_lamp(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    if already_has_lamp(ws):
        print(f"SKIP already has Equip_MinerLamp: {path}")
        return
    for row in LAMP_ROWS:
        ws.append(row)
    wb.save(path)
    print(f"OK appended 5 rows: {path}")


def main() -> None:
    targets = [
        ROOT / "Excel" / XLSX_NAME,
        ROOT / "Mode2" / "Excel" / XLSX_NAME,
    ]
    for path in targets:
        if not path.is_file():
            raise FileNotFoundError(path)
        append_lamp(path)


if __name__ == "__main__":
    main()
