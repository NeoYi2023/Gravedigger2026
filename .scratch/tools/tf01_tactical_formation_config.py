# -*- coding: utf-8 -*-
"""TF-01: Combat_TacticalFormationConfig + SkillConfig.FormationId + MagicBook_Form_Wedge.

Excel three-row headers preserved (SPEC_04 §14.7). Bakes Mode1 + Mode2 CSV.
"""
from __future__ import annotations

import io
import math
import re
import uuid
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_MAX_DECIMAL_PLACES = 10

FORMATION_XLSX = "战斗_战术阵型配置表_Combat_TacticalFormationConfig.xlsx"
FORMATION_CSV = "Combat_TacticalFormationConfig.csv"
SKILL_XLSX = "战斗_技能配置表_Combat_SkillConfig.xlsx"
SKILL_CSV = "Combat_SkillConfig.csv"
BOOK_XLSX = "制造_魔法书配置表_Manufacture_MagicBookConfig.xlsx"
BOOK_CSV = "Manufacture_MagicBookConfig.csv"

FORMATION_ZH = [
    "阵型ID",
    "阵型名称",
    "阵型图标",
    "阵型介绍",
    "阵型技能ID",
    "最小激活人数",
    "最大入阵人数",
    "阵型预制体ID",
    "属性加成",
    "专属技能ID",
    "专属技能效果ID",
]
FORMATION_NOTE = [
    "主键；被 GrantFormationSkill / SkillConfig.FormationId 引用",
    "UI 展示",
    "运行时 Resources/UI/Formations/{IconAssetId}；缺图占位",
    "UI 只读；不作效果器",
    "FK → SkillConfig.SkillId；命中后写入 SoldierSkills",
    "≥1；上阵同技能分组 ≥ 此值才组阵",
    "≥ MinMemberCount；不超过 Pattern 槽位数",
    "逻辑名 → Prefabs/Formation/Patterns/{PrefabId}.prefab",
    "Stat=…|Mul=… overlay；空=无",
    "SkillId|SkillId 激活态 overlay；空=无",
    "可选 SkillEffectId|…；空=无",
]
FORMATION_EN = [
    "FormationId",
    "DisplayName",
    "IconAssetId",
    "Description",
    "FormationSkillId",
    "MinMemberCount",
    "MaxMemberCount",
    "PrefabId",
    "StatModifiers",
    "ExclusiveSkillIds",
    "ExclusiveSkillEffectIds",
]
FORMATION_ROW = [
    "Form_Wedge_01",
    "楔阵",
    "Form_Wedge_01",
    "拥有楔阵技能的已上阵士兵可组成楔形编队（标记技能，不驱动战斗）",
    "Skill_Form_Wedge",
    3,
    5,
    "FormationPattern_Wedge_01",
    "Stat=Strength|Mul=1.15",
    "",
    "",
]

SKILL_SAMPLE = {
    "SkillId": "Skill_Form_Wedge",
    "SkillLevel": 1,
    "FormationId": "Form_Wedge_01",
    "CooldownMode": "Mode2",
    "CastTarget": "Self",
    "ExtraActivationCondition": "",
    "DisplayName": "楔阵",
    "Description": "战术阵型标记技能；制造授予后用于组阵判定，不驱动战斗",
    "IconAssetId": "Skill_Form_Wedge",
    "SkillEffectId": "",
    "BaseCooldownSeconds": 0,
    "LossOfControlChanceBonus": 0,
    "EffectImplemented": 0,
}

BOOK_SAMPLE = {
    "MagicBookId": "MagicBook_Form_Wedge",
    "IsUnique": 0,
    "IsProbabilistic": 0,
    "EffectPhase": "SoldierManufacture",
    "EffectPayload": "GrantFormationSkill",
    "EffectParams": "FormationId=Form_Wedge_01",
    "IconAssetId": "MagicBook_Form_Wedge",
    "DisplayName": "楔阵",
    "Description": "装备后，Mode2 制造 Step2 命中时授予楔阵技能（写入 SoldierSkills）",
    "VisualStyleId": "",
    "VisualPriority": "",
    "VisualIntensityAdd": "",
}

META_XLSX = """fileFormatVersion: 2
guid: {guid}
DefaultImporter:
  externalObjects: {{}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""

META_CSV = """fileFormatVersion: 2
guid: {guid}
TextScriptImporter:
  externalObjects: {{}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def sanitize_embedded_float_noise(text: str) -> str:
    if not text or "." not in text:
        return text

    def repl(match: re.Match[str]) -> str:
        token = match.group(0)
        frac = token.split(".", 1)[1]
        looks_noisy = len(frac) > 10 or "000000" in frac or "999999" in frac
        if not looks_noisy:
            return token
        try:
            number = float(token)
        except ValueError:
            return token
        return format_numeric_for_csv(number)

    return re.sub(r"-?\d+\.\d+", repl, text)


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    if isinstance(value, Decimal):
        return format_numeric_for_csv(float(value))
    return sanitize_embedded_float_noise(str(value))


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def bake(xlsx: Path, csv_out: Path) -> None:
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_csv_text(v) for v in row])
    wb.close()
    if not rows:
        raise SystemExit(f"empty workbook: {xlsx}")
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = None
    for i, row in enumerate(rows[:3]):
        if is_english_header(row):
            header_i = i
            break
    if header_i is None:
        raise SystemExit(f"no English header: {xlsx}")
    lines: list[str] = []
    for r_i, row in enumerate(rows[header_i:]):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    csv_out.parent.mkdir(parents=True, exist_ok=True)
    csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"Baked {csv_out} data_rows={len(lines) - 1}")


def write_meta(path: Path, template: str) -> None:
    if path.exists():
        return
    path.write_text(template.format(guid=uuid.uuid4().hex), encoding="utf-8")


def write_formation_xlsx(xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(FORMATION_ZH)
    ws.append(FORMATION_NOTE)
    ws.append(FORMATION_EN)
    ws.append(FORMATION_ROW)
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx)
    write_meta(xlsx.with_suffix(xlsx.suffix + ".meta"), META_XLSX)
    print(f"Wrote {xlsx}")


def header_map(ws) -> dict[str, int]:
    for r in range(1, 4):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        texts = ["" if v is None else str(v).strip() for v in values]
        if is_english_header(texts):
            return {name: i + 1 for i, name in enumerate(texts) if name}
    raise SystemExit("no English header in sheet")


def ensure_formation_id_column(ws) -> None:
    cols = header_map(ws)
    if "FormationId" in cols:
        return
    if "SkillLevel" not in cols:
        raise SystemExit("SkillConfig missing SkillLevel")
    insert_at = cols["SkillLevel"] + 1
    ws.insert_cols(insert_at)
    ws.cell(1, insert_at, "阵型ID")
    ws.cell(2, insert_at, "可选；空=普通士兵技能；非空=阵型标记技能 FK→TacticalFormationConfig")
    ws.cell(3, insert_at, "FormationId")
    print(f"  inserted FormationId at col {insert_at}")


def append_skill_row(ws) -> None:
    cols = header_map(ws)
    skill_col = cols["SkillId"]
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, skill_col).value
        if sid is not None and str(sid).strip() == SKILL_SAMPLE["SkillId"]:
            print(f"  Skill_Form_Wedge already present row {r}")
            return
    row_idx = ws.max_row + 1
    for name, value in SKILL_SAMPLE.items():
        if name not in cols:
            continue
        ws.cell(row_idx, cols[name], value)
    print(f"  appended Skill_Form_Wedge at row {row_idx}")


def append_book_row(ws) -> None:
    cols = header_map(ws)
    id_col = cols["MagicBookId"]
    for r in range(4, ws.max_row + 1):
        bid = ws.cell(r, id_col).value
        if bid is not None and str(bid).strip() == BOOK_SAMPLE["MagicBookId"]:
            print(f"  MagicBook_Form_Wedge already present row {r}")
            return
    row_idx = ws.max_row + 1
    for name, value in BOOK_SAMPLE.items():
        if name not in cols:
            continue
        ws.cell(row_idx, cols[name], value)
    print(f"  appended MagicBook_Form_Wedge at row {row_idx}")


def patch_skill(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    ensure_formation_id_column(ws)
    append_skill_row(ws)
    wb.save(xlsx)
    print(f"Patched {xlsx.name}")


def patch_book(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    append_book_row(ws)
    wb.save(xlsx)
    print(f"Patched {xlsx.name}")


def main() -> None:
    pairs = [
        (ROOT / "Excel", ROOT / "Csv"),
        (ROOT / "Mode2" / "Excel", ROOT / "Mode2" / "Csv"),
    ]
    for excel_dir, csv_dir in pairs:
        formation_xlsx = excel_dir / FORMATION_XLSX
        write_formation_xlsx(formation_xlsx)
        bake(formation_xlsx, csv_dir / FORMATION_CSV)
        write_meta((csv_dir / FORMATION_CSV).with_suffix(".csv.meta"), META_CSV)

        skill_xlsx = excel_dir / SKILL_XLSX
        patch_skill(skill_xlsx)
        bake(skill_xlsx, csv_dir / SKILL_CSV)

        book_xlsx = excel_dir / BOOK_XLSX
        patch_book(book_xlsx)
        bake(book_xlsx, csv_dir / BOOK_CSV)


if __name__ == "__main__":
    main()
