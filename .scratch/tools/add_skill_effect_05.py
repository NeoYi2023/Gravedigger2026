# -*- coding: utf-8 -*-
"""Append SkillEffect_05_1..5 to Mode2 Combat_SkillEffectConfig and bake CSV."""
from __future__ import annotations

import io
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables" / "Mode2"
XLSX = ROOT / "Excel" / "战斗_技能效果配置表_Combat_SkillEffectConfig.xlsx"
CSV_OUT = ROOT / "Csv" / "Combat_SkillEffectConfig.csv"

NEW_ROWS = [
    ("SkillEffect_05_1", "坚挺 Lv1：Self、敌人的本次攻击导致Self死亡；HP 强制变为 1，无敌 1 秒；BaseCD=60s"),
    ("SkillEffect_05_2", "坚挺 Lv2：同触发；无敌 2 秒"),
    ("SkillEffect_05_3", "坚挺 Lv3：同触发；无敌 3 秒"),
    ("SkillEffect_05_4", "坚挺 Lv4：同触发；无敌 4 秒"),
    ("SkillEffect_05_5", "坚挺 Lv5：同触发；无敌 5 秒"),
]

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
MAX_SCAN = 3


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def find_header_index(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows[:MAX_SCAN]):
        if is_english_header(row):
            return i
    raise ValueError("no English header in first 3 rows")


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def bake_csv() -> None:
    with open(XLSX, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_text(v) for v in row])
    wb.close()

    if not rows:
        raise ValueError("empty sheet")

    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))

    header_i = find_header_index(rows)
    export = rows[header_i:]
    lines: list[str] = []
    for r_i, row in enumerate(export):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))

    text = "\n".join(lines) + ("\n" if lines else "")
    CSV_OUT.write_text(text, encoding="utf-8", newline="\n")
    print(f"Baked {CSV_OUT.name}: {len(lines)} data rows (header at doc row {header_i + 1})")


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb.active
    existing_ids = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)}
    for sid, note in NEW_ROWS:
        if sid in existing_ids:
            print(f"SKIP exists {sid}")
            continue
        ws.append([sid, note])
        print(f"APPEND {sid}")
    wb.save(XLSX)
    print(f"Saved {XLSX.name}")
    bake_csv()


if __name__ == "__main__":
    main()
