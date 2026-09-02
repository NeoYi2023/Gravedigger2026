# -*- coding: utf-8 -*-
"""Append parallel-formation sample Form_Wedge_02 (+ skill + book) and bake CSV.

Mode1 + Mode2 Excel → Bake (SPEC_04 §14.7). Does not rewrite existing rows.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tf01_tactical_formation_config import (
    BOOK_CSV,
    BOOK_XLSX,
    FORMATION_CSV,
    FORMATION_XLSX,
    ROOT,
    SKILL_CSV,
    SKILL_XLSX,
    bake,
    header_map,
)

FORMATION_ROW = {
    "FormationId": "Form_Wedge_02",
    "DisplayName": "平行阵",
    "IconAssetId": "Form_Wedge_02",
    "Description": "拥有平行阵技能的已上阵士兵可组成两排平行编队（标记技能，不驱动战斗）",
    "FormationSkillId": "Skill_Form_Wedge_02",
    "MinMemberCount": 3,
    "MaxMemberCount": 10,
    "PrefabId": "FormationPattern_Wedge_02",
    "StatModifiers": "Stat=Strength|Mul=1.15",
    "ExclusiveSkillIds": "",
    "ExclusiveSkillEffectIds": "",
}

SKILL_ROW = {
    "SkillId": "Skill_Form_Wedge_02",
    "SkillLevel": 1,
    "FormationId": "Form_Wedge_02",
    "CooldownMode": "Mode2",
    "CastTarget": "Self",
    "ExtraActivationCondition": "",
    "DisplayName": "平行阵",
    "Description": "战术阵型标记技能；制造授予后用于组阵判定，不驱动战斗",
    "IconAssetId": "Skill_Form_Wedge_02",
    "SkillEffectId": "",
    "BaseCooldownSeconds": 0,
    "LossOfControlChanceBonus": 0,
    "EffectImplemented": 0,
}

BOOK_ROW = {
    "MagicBookId": "MagicBook_Form_Wedge_02",
    "IsUnique": 0,
    "IsProbabilistic": 0,
    "EffectPhase": "SoldierManufacture",
    "EffectPayload": "GrantFormationSkill",
    "EffectParams": "FormationId=Form_Wedge_02",
    "IconAssetId": "MagicBook_Form_Wedge_02",
    "DisplayName": "平行阵",
    "Description": "装备后，Mode2 制造 Step2 命中时授予平行阵技能（写入 SoldierSkills）",
    "VisualStyleId": "",
    "VisualPriority": "",
    "VisualIntensityAdd": "",
}


def append_by_pk(ws, pk_col_name: str, pk_value: str, row: dict) -> None:
    cols = header_map(ws)
    pk_col = cols[pk_col_name]
    for r in range(4, ws.max_row + 1):
        cell = ws.cell(r, pk_col).value
        if cell is not None and str(cell).strip() == pk_value:
            print(f"  {pk_value} already present row {r}")
            return
    row_idx = ws.max_row + 1
    for name, value in row.items():
        if name not in cols:
            continue
        ws.cell(row_idx, cols[name], value)
    print(f"  appended {pk_value} at row {row_idx}")


def patch_xlsx(xlsx: Path, pk_col: str, pk_value: str, row: dict) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    append_by_pk(ws, pk_col, pk_value, row)
    wb.save(xlsx)
    print(f"Patched {xlsx.name}")


def main() -> None:
    pairs = [
        (ROOT / "Excel", ROOT / "Csv"),
        (ROOT / "Mode2" / "Excel", ROOT / "Mode2" / "Csv"),
    ]
    for excel_dir, csv_dir in pairs:
        formation_xlsx = excel_dir / FORMATION_XLSX
        patch_xlsx(formation_xlsx, "FormationId", FORMATION_ROW["FormationId"], FORMATION_ROW)
        bake(formation_xlsx, csv_dir / FORMATION_CSV)

        skill_xlsx = excel_dir / SKILL_XLSX
        patch_xlsx(skill_xlsx, "SkillId", SKILL_ROW["SkillId"], SKILL_ROW)
        bake(skill_xlsx, csv_dir / SKILL_CSV)

        book_xlsx = excel_dir / BOOK_XLSX
        patch_xlsx(book_xlsx, "MagicBookId", BOOK_ROW["MagicBookId"], BOOK_ROW)
        bake(book_xlsx, csv_dir / BOOK_CSV)


if __name__ == "__main__":
    main()
