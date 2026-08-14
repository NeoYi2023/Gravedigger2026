# -*- coding: utf-8 -*-
"""Insert BaseClass column into Manufacture_ClassConfig Excel + rewrite CSV."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"e:\Work\Cursor\Gravedigger2026\Gravedigger2026\Gravedigger2026\Assets\ConfigTables")

WARRIOR = "战士"
ARCHER = "射手"
MAGE = "法师"
THIEF = "盗贼"

ARCHER_IDS = {
    "Class_Ranger",
    "Class_BombMaster",
    "Class_Longbowman",
}
MAGE_IDS = {
    "Class_Priest",
    "Class_Warlock",
    "Class_IceMage",
    "Class_FireMage",
    "Class_DarkMage",
}

DESC = (
    "CSV 中文：战士|射手|法师|盗贼 → 枚举；空/非法→Unspecified；"
    "预留后续魔法书等条件；不参与命名/外观/PrimaryStat/战斗；不烘进实例"
)


def map_base(class_id: str) -> str:
    if class_id.startswith("Class_Rogue") or class_id == "Class_Shadowblade":
        return THIEF
    if class_id.startswith("Class_Archer") or class_id in ARCHER_IDS:
        return ARCHER
    if class_id.startswith("Class_Mage") or class_id in MAGE_IDS:
        return MAGE
    return WARRIOR


def patch_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active

    # Row 3 = English headers (SPEC_04 §14 three-row)
    en_row = 3
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(en_row, c).value or "").strip()
        if v:
            headers[v] = c

    if "ClassId" not in headers or "ClassName" not in headers:
        raise SystemExit(f"Bad headers in {path}: {headers}")

    if "BaseClass" in headers:
        base_col = headers["BaseClass"]
        print(f"  BaseClass already at col {base_col}")
    else:
        insert_at = headers["ClassName"] + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "基础职业"
        ws.cell(2, insert_at).value = DESC
        ws.cell(3, insert_at).value = "BaseClass"
        base_col = insert_at
        print(f"  inserted BaseClass at col {base_col}")

    class_id_col = headers["ClassId"]
    if "BaseClass" not in headers and class_id_col >= base_col:
        # ClassId is before ClassName, unaffected
        pass
    # Re-resolve ClassId after possible insert (ClassId is col 1, always before insert)
    class_id_col = 1
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(en_row, c).value or "").strip() == "ClassId":
            class_id_col = c
            break

    filled = 0
    for r in range(4, ws.max_row + 1):
        cid = ws.cell(r, class_id_col).value
        if cid is None:
            continue
        cid = str(cid).strip()
        if not cid.startswith("Class_"):
            continue
        ws.cell(r, base_col).value = map_base(cid)
        filled += 1

    wb.save(path)
    print(f"  filled {filled} rows in {path.name}")


def rewrite_csv(path: Path) -> None:
    text = path.read_text(encoding="utf-8-sig")
    lines = text.splitlines()
    if not lines:
        raise SystemExit(f"Empty CSV {path}")
    header = lines[0].split(",")
    if "BaseClass" in header:
        idx = header.index("BaseClass")
        out_header = header
    else:
        if "ClassName" not in header:
            raise SystemExit(f"No ClassName in {path}")
        name_i = header.index("ClassName")
        out_header = header[: name_i + 1] + ["BaseClass"] + header[name_i + 1 :]
        idx = name_i + 1

    out_lines = [",".join(out_header)]
    id_i = out_header.index("ClassId")
    for line in lines[1:]:
        if not line.strip():
            continue
        parts = line.split(",")
        if "BaseClass" not in header:
            parts = parts[:idx] + [""] + parts[idx:]
        while len(parts) < len(out_header):
            parts.append("")
        cid = parts[id_i].strip()
        parts[idx] = map_base(cid)
        out_lines.append(",".join(parts[: len(out_header)]))

    path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    print(f"  wrote CSV {path.name} rows={len(out_lines) - 1}")


def main() -> None:
    for rel in (
        "Excel/制造_职业配置表_Manufacture_ClassConfig.xlsx",
        "Mode2/Excel/制造_职业配置表_Manufacture_ClassConfig.xlsx",
    ):
        p = ROOT / rel
        print(f"=== Excel {p.name}")
        patch_xlsx(p)

    for rel in (
        "Csv/Manufacture_ClassConfig.csv",
        "Mode2/Csv/Manufacture_ClassConfig.csv",
    ):
        p = ROOT / rel
        print(f"=== CSV {p.name}")
        rewrite_csv(p)


if __name__ == "__main__":
    main()
