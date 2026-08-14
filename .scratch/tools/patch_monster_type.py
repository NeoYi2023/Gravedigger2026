# -*- coding: utf-8 -*-
"""Insert MonsterType column into Defend_MonsterConfig Excel + rewrite CSV."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"e:\Work\Cursor\Gravedigger2026\Gravedigger2026\Gravedigger2026\Assets\ConfigTables")
XLSX_NAME = "防守_怪物配置表_Defend_MonsterConfig.xlsx"
CSV_NAME = "Defend_MonsterConfig.csv"

DESC = (
    "1=普通(Normal) / 2=精英(Elite) / 3=BOSS(Boss)；"
    "缺列或空→1；非法→加载失败；异于 PushMapSpawnConfig.IsBoss；本批不驱动技能"
)


def monster_type_for(monster_id: str) -> int:
    return 3 if monster_id == "Monster_11" else 1


def patch_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    en_row = 3
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(en_row, c).value or "").strip()
        if v:
            headers[v] = c

    if "AttackMode" not in headers or "MonsterId" not in headers:
        raise SystemExit(f"Bad headers in {path}: {list(headers)}")

    if "MonsterType" in headers:
        type_col = headers["MonsterType"]
        print(f"  MonsterType already at col {type_col}")
    else:
        insert_at = headers["AttackMode"] + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "怪物类型"
        ws.cell(2, insert_at).value = DESC
        ws.cell(3, insert_at).value = "MonsterType"
        type_col = insert_at
        print(f"  inserted MonsterType at col {type_col}")

    # Re-resolve MonsterId after possible insert
    monster_id_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(en_row, c).value or "").strip() == "MonsterId":
            monster_id_col = c
            break
    if monster_id_col is None:
        raise SystemExit(f"MonsterId missing after patch in {path}")

    for r in range(4, ws.max_row + 1):
        mid = str(ws.cell(r, monster_id_col).value or "").strip()
        if not mid:
            continue
        ws.cell(r, type_col).value = monster_type_for(mid)

    wb.save(path)
    print(f"OK xlsx {path}")


def bake_csv(xlsx_path: Path, csv_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    en_row = 3
    headers = []
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(en_row, c).value or "").strip()
        headers.append(v)

    rows = []
    for r in range(4, ws.max_row + 1):
        mid = str(ws.cell(r, 1).value or "").strip()
        if not mid:
            continue
        row = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c).value
            if cell is None:
                row.append("")
            else:
                row.append(str(cell))
        rows.append(row)

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"OK csv  {csv_path} ({len(rows)} rows)")


def main() -> None:
    pairs = [
        (ROOT / "Excel" / XLSX_NAME, ROOT / "Csv" / CSV_NAME),
        (ROOT / "Mode2" / "Excel" / XLSX_NAME, ROOT / "Mode2" / "Csv" / CSV_NAME),
    ]
    for xlsx, csv_path in pairs:
        if not xlsx.exists():
            raise SystemExit(f"Missing {xlsx}")
        patch_xlsx(xlsx)
        bake_csv(xlsx, csv_path)


if __name__ == "__main__":
    main()
