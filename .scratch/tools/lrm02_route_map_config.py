# -*- coding: utf-8 -*-
"""LRM-02: add RouteMapAssetId + MapPosX/Y to Level tables (Mode1+Mode2), then bake CSV."""
from __future__ import annotations

import io
import math
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
MODE_ROOTS = [ROOT, ROOT / "Mode2"]

OP_XLSX = "关卡_关卡运作表_Level_LevelOperationConfig.xlsx"
SUB_XLSX = "关卡_子关卡表_Level_SubLevelConfig.xlsx"

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
MAX_SCAN = 3
_MAX_DECIMAL_PLACES = 10

# Mode2 demo: hang SubLevel_001 on Level_01..03; placeholder card centers (1450-wide map).
MODE2_MAP_ID = "SubLevel_001"
MODE2_POS = {
    # Level_01 — bottom → top along path landmarks
    "Opt_L01_S1_Shop": (725, 280),
    "Opt_L01_S2_Dig_A": (480, 720),
    "Opt_L01_S2_Dig_B": (980, 720),
    "Opt_L01_S3_AutoManufacture": (725, 1180),
    "Opt_L01_S4_UpgradeManufacture": (725, 1680),
    "Opt_L01_S5_PushMap": (520, 2200),
    "Opt_SE_Demo_01": (980, 2200),
    # Level_02
    "Opt_L02_S1_Shop": (725, 280),
    "Opt_L02_S2_Dig": (725, 720),
    "Opt_L02_S3_AutoManufacture": (725, 1180),
    "Opt_L02_S4_UpgradeManufacture": (725, 1680),
    "Opt_L02_S5_PushMap": (725, 2200),
    # Level_03
    "Opt_L03_S1_Shop": (725, 280),
    "Opt_L03_S2_Dig": (725, 720),
    "Opt_L03_S3_AutoManufacture": (725, 1180),
    "Opt_L03_S4_UpgradeManufacture": (725, 1680),
    "Opt_L03_S5_PushMap": (725, 2200),
}


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def sanitize_embedded_float_noise(text: str) -> str:
    if not text or "." not in text:
        return text

    def repl(match: re.Match[str]) -> str:
        token = match.group(0)
        frac = token.split(".", 1)[1]
        looks_noisy = len(frac) > 10 or "000000" in frac or "999999" in frac
        if not looks_noisy:
            return token
        try:
            number = float(token)
        except ValueError:
            return token
        return format_numeric_for_csv(number)

    return re.sub(r"-?\d+\.\d+", repl, text)


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    if isinstance(value, Decimal):
        return format_numeric_for_csv(float(value))
    return sanitize_embedded_float_noise(str(value))


def excel_to_csv_base(stem: str) -> str:
    parts = stem.split("_")
    if len(parts) != 4:
        raise ValueError(f"bad name: {stem}")
    return f"{parts[2]}_{parts[3]}"


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def find_header_index(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows[:MAX_SCAN]):
        if is_english_header(row):
            return i
    raise ValueError("no English header in first 3 rows")


def ensure_column(ws, en_name: str, zh: str, note: str) -> int:
    headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    if en_name in headers:
        return headers.index(en_name) + 1
    col = len(headers) + 1
    # trim trailing Nones from headers length
    while headers and headers[-1] is None:
        headers.pop()
        col = len(headers) + 1
    # recompute: find last non-empty in row 3
    last = 0
    for cell in ws[3]:
        if cell.value is not None and str(cell.value).strip() != "":
            last = cell.column
    col = last + 1
    ws.cell(1, col, zh)
    ws.cell(2, col, note)
    ws.cell(3, col, en_name)
    return col


def patch_operation(xlsx: Path, fill_mode2_map: bool) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    col = ensure_column(
        ws,
        "RouteMapAssetId",
        "路线底图资源ID",
        "可选；文件名无扩展名；同 LevelId 取首个非空；空=无底图",
    )
    if fill_mode2_map:
        headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
        level_col = headers.index("LevelId") + 1
        for row in ws.iter_rows(min_row=4):
            level = row[level_col - 1].value
            if level in ("Level_01", "Level_02", "Level_03"):
                # only first row per level needs value; fill all for simplicity (same id)
                ws.cell(row[0].row, col, MODE2_MAP_ID)
    wb.save(xlsx)
    print(f"patched {xlsx}")


def patch_sublevel(xlsx: Path, fill_mode2_pos: bool) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    x_col = ensure_column(
        ws,
        "MapPosX",
        "地图坐标X",
        "选项卡中心；底图左下原点；展示宽1450 UI像素；空=0",
    )
    y_col = ensure_column(
        ws,
        "MapPosY",
        "地图坐标Y",
        "选项卡中心；Y向上；空=0",
    )
    if fill_mode2_pos:
        headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
        id_col = headers.index("GameplayOptionId") + 1
        for row in ws.iter_rows(min_row=4):
            oid = row[id_col - 1].value
            if oid in MODE2_POS:
                x, y = MODE2_POS[oid]
                ws.cell(row[0].row, x_col, x)
                ws.cell(row[0].row, y_col, y)
    wb.save(xlsx)
    print(f"patched {xlsx}")


def bake_one(cfg_root: Path, xlsx_name: str) -> None:
    excel_dir = cfg_root / "Excel"
    csv_dir = cfg_root / "Csv"
    xlsx = excel_dir / xlsx_name
    if not xlsx.is_file():
        print(f"skip bake missing {xlsx}")
        return
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [cell_to_csv_text(v) for v in row]
        rows.append(cells)
    wb.close()
    if not rows:
        return
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = find_header_index(rows)
    export = rows[header_i:]
    lines = []
    for r_i, row in enumerate(export):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(
            ",".join(
                ('"' + c.replace('"', '""') + '"' if any(ch in c for ch in [",", '"', "\n", "\r"]) else c)
                for c in (c or "" for c in row)
            )
        )
    text = "\n".join(lines) + ("\n" if lines else "")
    out = csv_dir / f"{excel_to_csv_base(xlsx.stem)}.csv"
    out.write_text(text, encoding="utf-8", newline="\n")
    print(f"baked {out}")


def main() -> None:
    for cfg in MODE_ROOTS:
        is_mode2 = cfg.name == "Mode2"
        op = cfg / "Excel" / OP_XLSX
        sub = cfg / "Excel" / SUB_XLSX
        if not op.is_file() or not sub.is_file():
            print(f"missing excel under {cfg}", file=sys.stderr)
            continue
        patch_operation(op, fill_mode2_map=is_mode2)
        patch_sublevel(sub, fill_mode2_pos=is_mode2)
        bake_one(cfg, OP_XLSX)
        bake_one(cfg, SUB_XLSX)


if __name__ == "__main__":
    main()
