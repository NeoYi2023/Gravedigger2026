# -*- coding: utf-8 -*-
"""LRM-06: drop SubLevel MapPosX/MapPosY (Mode1+Mode2 Excel), then bake CSV."""
from __future__ import annotations

import io
import math
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
MODE_ROOTS = [ROOT, ROOT / "Mode2"]
SUB_XLSX = "关卡_子关卡表_Level_SubLevelConfig.xlsx"
DROP_COLS = ("MapPosX", "MapPosY")

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
MAX_SCAN = 3
_MAX_DECIMAL_PLACES = 10


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def sanitize_embedded_float_noise(text: str) -> str:
    if not text or "." not in text:
        return text

    def repl(match: re.Match[str]) -> str:
        token = match.group(0)
        frac = token.split(".", 1)[1]
        looks_noisy = len(frac) > 10 or "000000" in frac or "999999" in frac
        if not looks_noisy:
            return token
        try:
            number = float(token)
        except ValueError:
            return token
        return format_numeric_for_csv(number)

    return re.sub(r"-?\d+\.\d+", repl, text)


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    if isinstance(value, Decimal):
        return format_numeric_for_csv(float(value))
    return sanitize_embedded_float_noise(str(value))


def excel_to_csv_base(stem: str) -> str:
    parts = stem.split("_")
    if len(parts) != 4:
        raise ValueError(f"bad name: {stem}")
    return f"{parts[2]}_{parts[3]}"


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def find_header_index(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows[:MAX_SCAN]):
        if is_english_header(row):
            return i
    raise ValueError("no English header in first 3 rows")


def drop_columns(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx)
    ws = wb.active
    headers = []
    for cell in ws[3]:
        headers.append(cell.value)
    while headers and headers[-1] is None:
        headers.pop()
    dropped = []
    for name in DROP_COLS:
        if name not in headers:
            continue
        col = headers.index(name) + 1
        ws.delete_cols(col)
        headers.pop(col - 1)
        dropped.append(name)
    if dropped:
        wb.save(xlsx)
    wb.close()
    return dropped


def bake_one(cfg_root: Path, xlsx_name: str) -> None:
    excel_dir = cfg_root / "Excel"
    csv_dir = cfg_root / "Csv"
    xlsx = excel_dir / xlsx_name
    if not xlsx.is_file():
        print(f"skip bake missing {xlsx}")
        return
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [cell_to_csv_text(v) for v in row]
        rows.append(cells)
    wb.close()
    if not rows:
        return
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = find_header_index(rows)
    export = rows[header_i:]
    lines = []
    for r_i, row in enumerate(export):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(
            ",".join(
                ('"' + c.replace('"', '""') + '"' if any(ch in c for ch in [",", '"', "\n", "\r"]) else c)
                for c in (c or "" for c in row)
            )
        )
    text = "\n".join(lines) + ("\n" if lines else "")
    out = csv_dir / f"{excel_to_csv_base(xlsx.stem)}.csv"
    out.write_text(text, encoding="utf-8", newline="\n")
    print(f"baked {out}")


def main() -> None:
    for cfg in MODE_ROOTS:
        sub = cfg / "Excel" / SUB_XLSX
        if not sub.is_file():
            print(f"missing excel {sub}", file=sys.stderr)
            continue
        dropped = drop_columns(sub)
        print(f"dropped {dropped or 'none'} from {sub}")
        bake_one(cfg, SUB_XLSX)


if __name__ == "__main__":
    main()
