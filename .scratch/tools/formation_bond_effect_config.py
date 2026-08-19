# -*- coding: utf-8 -*-
"""Append BondEffect_* placeholder rows to SkillEffectConfig (Mode1+Mode2) and bake CSV."""
from __future__ import annotations

import io
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")

MODE2_BOND_EFFECTS = [
    ("BondEffect_IronWall_1", "铜墙铁壁 I：战士 MaxHP +3%（展示；Handler 未接线）"),
    ("BondEffect_IronWall_2", "铜墙铁壁 II：战士 MaxHP +5%（展示；Handler 未接线）"),
    ("BondEffect_HumanLegion_1", "人类军团：展示用羁绊（Handler 未接线）"),
    ("BondEffect_FullArmy_1", "满编：展示用羁绊（Handler 未接线）"),
    ("BondEffect_StrStrength_1", "力之共鸣：展示用羁绊（Handler 未接线）"),
    ("BondEffect_PreciseClass_1", "近卫专精：展示用羁绊（Handler 未接线）"),
]

MODE1_BOND_EFFECTS = [
    ("BondEffect_Demo_1", "Mode1 演示羁绊：展示占位（Handler 未接线）"),
]


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-10)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.10f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    return str(value)


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def bake(xlsx: Path, csv_out: Path) -> None:
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_csv_text(v) for v in row])
    wb.close()
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = next(i for i, row in enumerate(rows[:3]) if is_english_header(row))
    lines: list[str] = []
    for r_i, row in enumerate(rows[header_i:]):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"Baked {csv_out}: data_rows={len(lines) - 1}")


def append_bond_effects(xlsx: Path, rows_to_add: list[tuple[str, str]], has_effect_columns: bool) -> int:
    wb = load_workbook(xlsx)
    ws = wb.active
    existing: set[str] = set()
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is not None:
            existing.add(str(sid).strip())

    added = 0
    next_row = ws.max_row + 1
    for sid, notes in rows_to_add:
        if sid in existing:
            print(f"SKIP {sid} (already exists)")
            continue
        ws.cell(next_row, 1, sid)
        ws.cell(next_row, 2, notes)
        if has_effect_columns:
            ws.cell(next_row, 3, None)
            ws.cell(next_row, 4, None)
            ws.cell(next_row, 5, None)
        added += 1
        print(f"ADD {sid}")
        next_row += 1

    wb.save(xlsx)
    print(f"Saved {xlsx.name}: added {added} rows")
    return added


def main() -> None:
    mode2_xlsx = ROOT / "Mode2" / "Excel" / "战斗_技能效果配置表_Combat_SkillEffectConfig.xlsx"
    mode2_csv = ROOT / "Mode2" / "Csv" / "Combat_SkillEffectConfig.csv"
    append_bond_effects(mode2_xlsx, MODE2_BOND_EFFECTS, has_effect_columns=True)
    bake(mode2_xlsx, mode2_csv)

    mode1_xlsx = ROOT / "Excel" / "战斗_技能效果配置表_Combat_SkillEffectConfig.xlsx"
    mode1_csv = ROOT / "Csv" / "Combat_SkillEffectConfig.csv"
    append_bond_effects(mode1_xlsx, MODE1_BOND_EFFECTS, has_effect_columns=False)
    bake(mode1_xlsx, mode1_csv)


if __name__ == "__main__":
    main()
