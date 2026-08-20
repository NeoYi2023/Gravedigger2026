# -*- coding: utf-8 -*-
"""Assemble App_1/2/3 warrior Prefabs + Mode2 BodyAppearanceConfig + catalogs."""
import re
import uuid
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"e:\Work\Cursor\Gravedigger2026\Gravedigger2026\Gravedigger2026")
ART = ROOT / "Assets/Art/Characters/Appearances"
PREFAB_DIR = ROOT / "Assets/Prefabs/Defend/Warriors"
CSV_PATH = ROOT / "Assets/ConfigTables/Mode2/Csv/Manufacture_BodyAppearanceConfig.csv"
XLSX_PATH = (
    ROOT
    / "Assets/ConfigTables/Mode2/Excel/制造_躯体外观配置表_Manufacture_BodyAppearanceConfig.xlsx"
)
DEFEND_CATALOG = ROOT / "Assets/Settings/Defend/DefendPrefabCatalog.asset"
UM_CATALOG = ROOT / "Assets/Settings/UpgradeManufacture/UpgradeManufacturePrefabCatalog.asset"
TEMPLATE = (PREFAB_DIR / "App_0_02.prefab").read_text(encoding="utf-8")

SUFFIX_META = {
    "02": (3, "近卫", "高防御战士"),
    "03": (3, "狂战士", "高伤害"),
    "12": (3, "炸弹师", "范围爆炸"),
    "13": (3, "长弓手", "直线贯穿"),
    "22": (3, "冰法", "减速冰冻"),
    "23": (3, "火法", "范围持续伤害"),
    "32": (3, "格斗师", "专供精英敌人"),
    "33": (3, "影刃", "瞬移后排"),
}
RACE = {"1": "Race_Human", "2": "Race_Elf", "3": "Race_Orc"}
ROOT_GO = "3927253789114556784"

IDS = []
for race_digit in ("1", "2", "3"):
    for suf in ("02", "03", "12", "13", "22", "23", "32", "33"):
        aid = f"App_{race_digit}_{suf}"
        IDS.append(aid)


def row_for(aid):
    parts = aid.split("_")
    race = RACE[parts[1]]
    suf = parts[2]
    level, affinity, desc = SUFFIX_META[suf]
    return [aid, level, race, affinity, desc, 1, 0.16, 0.6, 0.6, 0]


def read_guid(meta_path):
    text = Path(meta_path).read_text(encoding="utf-8")
    m = re.search(r"^guid: ([a-f0-9]+)", text, re.M)
    return m.group(1) if m else None


def find_controller(art_dir: Path):
    for p in art_dir.rglob("*.controller"):
        return p, read_guid(str(p) + ".meta")
    return None, None


def find_idle_sprite(art_dir: Path):
    idle = art_dir / "Idle.png"
    meta = idle.with_suffix(".png.meta")
    if not idle.exists():
        return None, None
    guid = read_guid(meta)
    text = meta.read_text(encoding="utf-8")
    m = re.search(r"name: Idle_2_0\n.*?internalID: (-?\d+)", text, re.S)
    if not m:
        m = re.search(r"name: Idle_0_0\n.*?internalID: (-?\d+)", text, re.S)
    if not m:
        m = re.search(r"internalID: (-?\d+)", text)
    return guid, m.group(1) if m else None


def append_catalog(path: Path, entries):
    text = path.read_text(encoding="utf-8")
    text = re.sub(
        r"\n  - AppearanceId: App_[123]_[^\n]+\n    Prefab: \{[^\n]+\}",
        "",
        text,
    )
    block = "".join(
        f"\n  - AppearanceId: {aid}\n"
        f"    Prefab: {{fileID: {ROOT_GO}, guid: {guid}, type: 3}}"
        for aid, guid in entries
    )
    if "\n  _monsterModels:" in text:
        text = text.replace("\n  _monsterModels:", block + "\n  _monsterModels:", 1)
    else:
        m = re.search(r"(  - AppearanceId: App_5_51\n    Prefab: \{[^\n]+\})", text)
        if not m:
            raise RuntimeError(f"cannot find App_5_51 anchor in {path}")
        text = text[: m.end()] + block + text[m.end() :]
    path.write_text(text, encoding="utf-8")
    print("catalog updated", path.name)


def main():
    print("target ids", len(IDS))

    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    for r in range(ws.max_row, 3, -1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and re.match(r"^App_[123]_", v):
            ws.delete_rows(r, 1)

    last = 3
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value:
            last = r
    start = last + 1
    for i, aid in enumerate(IDS):
        for c, val in enumerate(row_for(aid), 1):
            ws.cell(start + i, c).value = val
    wb.save(XLSX_PATH)
    print("excel updated from row", start)

    undead = (
        "AppearanceId,AppearanceLevel,RaceId,ClassAffinity,Description,"
        "IsFallback,BodyRadius,PushCoefficient,RepulsionScale,FacingYawFlip\n"
        "App_0_00,1,Race_Undead,枯骨战士,基础兵种,1,0.13,0.4,0.4,0\n"
        "App_0_10,1,Race_Undead,枯骨射手,基础兵种,1,0.13,0.4,0.4,0\n"
        "App_0_20,1,Race_Undead,枯骨法师,基础兵种,1,0.13,0.4,0.4,0\n"
        "App_0_30,1,Race_Undead,枯骨刺客,基础兵种,1,0.13,0.4,0.4,0\n"
        "App_0_01,2,Race_Undead,战士,2级战士,1,0.16,0.5,0.5,0\n"
        "App_0_11,2,Race_Undead,射手,2级射手,1,0.16,0.5,0.5,0\n"
        "App_0_21,2,Race_Undead,法师,2级法师,1,0.16,0.5,0.5,0\n"
        "App_0_31,2,Race_Undead,刺客,2级刺客,1,0.16,0.5,0.5,0\n"
        "App_0_02,3,Race_Undead,近卫,高防御战士,1,0.16,0.6,0.6,0\n"
        "App_0_12,3,Race_Undead,炸弹师,范围爆炸,1,0.16,0.6,0.6,0\n"
        "App_0_22,3,Race_Undead,冰法,减速冰冻,1,0.16,0.6,0.6,0\n"
        "App_0_32,3,Race_Undead,格斗师,专供精英敌人,1,0.16,0.6,0.6,0\n"
        "App_0_03,3,Race_Undead,狂战士,高伤害,1,0.16,0.6,0.6,0\n"
        "App_0_13,3,Race_Undead,长弓手,直线贯穿,1,0.16,0.6,0.6,0\n"
        "App_0_23,3,Race_Undead,火法,范围持续伤害,1,0.16,0.6,0.6,0\n"
        "App_0_33,3,Race_Undead,影刃,瞬移后排,1,0.16,0.6,0.6,0\n"
        "App_4_41,4,Race_Angel,圣骑士,范围防御,1,0.16,0.7,0.7,0\n"
        "App_5_51,4,Race_Demon,暗黑法师,致命魔法,1,0.16,0.7,0.7,0\n"
    )
    lines = [undead.rstrip()]
    for aid in IDS:
        lines.append(",".join(str(x) for x in row_for(aid)))
    CSV_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    print("csv written")

    created = []
    for aid in IDS:
        art_dir = ART / aid
        if not art_dir.is_dir():
            print("MISSING ART", aid)
            continue
        ctrl_path, ctrl_guid = find_controller(art_dir)
        sprite_guid, sprite_id = find_idle_sprite(art_dir)
        if not ctrl_guid or not sprite_guid or not sprite_id:
            print("MISSING REFS", aid, ctrl_guid, sprite_guid, sprite_id)
            continue

        ctext = ctrl_path.read_text(encoding="utf-8")
        ctext2, n = re.subn(
            r"(- m_Name: DirIndex\n    m_Type: 3\n    m_DefaultFloat: 0\n    m_DefaultInt: )0",
            r"\g<1>2",
            ctext,
            count=1,
        )
        if n:
            ctrl_path.write_text(ctext2, encoding="utf-8")

        prefab = TEMPLATE.replace("m_Name: App_0_02", f"m_Name: {aid}")
        prefab = re.sub(
            r"m_Sprite: \{fileID: -?\d+, guid: [a-f0-9]+, type: 3\}",
            f"m_Sprite: {{fileID: {sprite_id}, guid: {sprite_guid}, type: 3}}",
            prefab,
            count=1,
        )
        prefab = re.sub(
            r"m_Controller: \{fileID: 9100000, guid: [a-f0-9]+, type: 2\}",
            f"m_Controller: {{fileID: 9100000, guid: {ctrl_guid}, type: 2}}",
            prefab,
            count=1,
        )
        out = PREFAB_DIR / f"{aid}.prefab"
        out.write_text(prefab, encoding="utf-8")
        meta_guid = uuid.uuid4().hex
        (PREFAB_DIR / f"{aid}.prefab.meta").write_text(
            "fileFormatVersion: 2\n"
            f"guid: {meta_guid}\n"
            "PrefabImporter:\n"
            "  externalObjects: {}\n"
            "  userData: \n"
            "  assetBundleName: \n"
            "  assetBundleVariant: \n",
            encoding="utf-8",
        )
        created.append((aid, meta_guid))
        print("prefab", aid)

    print("created", len(created))
    append_catalog(DEFEND_CATALOG, created)
    append_catalog(UM_CATALOG, created)

    missing = [aid for aid in IDS if not (PREFAB_DIR / f"{aid}.prefab").exists()]
    print("missing prefabs", missing)
    print("DONE")


if __name__ == "__main__":
    main()
