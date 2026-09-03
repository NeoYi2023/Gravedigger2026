# -*- coding: utf-8 -*-
"""v0.84.02: Level_DifficultyConfig + LevelOperationConfig.DifficultyId.

Creates Mode1+Mode2 Difficulty Excel/CSV; inserts DifficultyId after LevelId
on Level_LevelOperationConfig (fill Diff_Normal). SPEC_04 §14.7 three-row headers.
"""
from __future__ import annotations

import re
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")

DIFF_XLSX = "关卡_关卡难度表_Level_DifficultyConfig.xlsx"
DIFF_CSV = "Level_DifficultyConfig.csv"
OP_XLSX = "关卡_关卡运作表_Level_LevelOperationConfig.xlsx"
OP_CSV = "Level_LevelOperationConfig.csv"

DIFF_ZH = ["难度ID", "显示名", "解锁所需难度ID", "文字介绍", "难度通关奖励"]
DIFF_NOTE = [
    "主键；样例 Diff_Normal/Hard/Hell",
    "UI 栏标题",
    "空=初始解锁；填本表 DifficultyId=等通关；找不到=不可解锁",
    "悬停/说明文案",
    "ItemId;Count|…；经 §9.5a；空=无；首次难度通关一次性发放",
]
DIFF_EN = [
    "DifficultyId",
    "DisplayName",
    "UnlockRequireDifficultyId",
    "Description",
    "ClearReward",
]
DIFF_ROWS = [
    [
        "Diff_Normal",
        "普通",
        "",
        "普通难度：适合熟悉关卡流程与基础玩法。",
        "",
    ],
    [
        "Diff_Hard",
        "困难",
        "Diff_Normal",
        "困难：通关普通难度全部关卡后解锁（Demo 接线后置）。",
        "Spirit;50",
    ],
    [
        "Diff_Hell",
        "地狱",
        "Diff_Hard",
        "地狱：通关困难难度全部关卡后解锁（Demo 接线后置）。",
        "Spirit;100",
    ],
]

META_XLSX = """fileFormatVersion: 2
guid: {guid}
DefaultImporter:
  externalObjects: {{}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""

META_CSV = """fileFormatVersion: 2
guid: {guid}
TextScriptImporter:
  externalObjects: {{}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if not (cell or "").strip():
            continue
        saw = True
        if not EN_COL.match(cell.strip()):
            return False
    return saw


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def write_meta(path: Path, template: str) -> None:
    if path.exists():
        return
    path.write_text(template.format(guid=uuid.uuid4().hex), encoding="utf-8")


def write_difficulty_xlsx(xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(DIFF_ZH)
    ws.append(DIFF_NOTE)
    ws.append(DIFF_EN)
    for row in DIFF_ROWS:
        ws.append(row)
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx)
    write_meta(xlsx.with_suffix(xlsx.suffix + ".meta"), META_XLSX)
    print(f"Wrote {xlsx}")


def bake_xlsx_to_csv(xlsx: Path, csv_out: Path) -> None:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    wb.close()
    if not rows:
        raise SystemExit(f"empty: {xlsx}")
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    # Trim trailing empty columns
    while max_cols > 0 and all(not (r[max_cols - 1] or "").strip() for r in rows):
        max_cols -= 1
        for r in rows:
            if len(r) > max_cols:
                del r[max_cols:]
    header_i = None
    for i, row in enumerate(rows[:3]):
        if is_english_header(row):
            header_i = i
            break
    if header_i is None:
        raise SystemExit(f"no English header: {xlsx}")
    lines: list[str] = []
    for r_i, row in enumerate(rows[header_i:]):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    csv_out.parent.mkdir(parents=True, exist_ok=True)
    csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    write_meta(csv_out.with_suffix(csv_out.suffix + ".meta"), META_CSV)
    print(f"Baked {csv_out} data_rows={len(lines) - 1}")


def header_map(ws) -> tuple[int, dict[str, int]]:
    for r in range(1, 4):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        texts = ["" if v is None else str(v).strip() for v in values]
        if is_english_header(texts):
            return r, {name: i + 1 for i, name in enumerate(texts) if name}
    raise SystemExit("no English header in LevelOperation sheet")


def ensure_difficulty_id_column(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    header_row, cols = header_map(ws)
    if "DifficultyId" in cols:
        # Ensure all data rows filled
        col = cols["DifficultyId"]
        filled = 0
        for r in range(header_row + 1, ws.max_row + 1):
            lid = ws.cell(r, cols["LevelId"]).value
            if lid is None or not str(lid).strip():
                continue
            cur = ws.cell(r, col).value
            if cur is None or not str(cur).strip():
                ws.cell(r, col, "Diff_Normal")
                filled += 1
        wb.save(xlsx)
        print(f"  DifficultyId present; filled empty={filled} in {xlsx.name}")
        return

    if "LevelId" not in cols:
        raise SystemExit(f"missing LevelId: {xlsx}")
    insert_at = cols["LevelId"] + 1
    ws.insert_cols(insert_at)
    # Doc rows 1–2 + English header
    ws.cell(1, insert_at, "难度ID")
    ws.cell(2, insert_at, "FK→DifficultyConfig；同 LevelId 多行须同值；空=不归属难度")
    ws.cell(header_row, insert_at, "DifficultyId")
    for r in range(header_row + 1, ws.max_row + 1):
        lid = ws.cell(r, cols["LevelId"]).value
        if lid is None or not str(lid).strip():
            continue
        ws.cell(r, insert_at, "Diff_Normal")
    wb.save(xlsx)
    print(f"  inserted DifficultyId at col {insert_at} in {xlsx.name}")


def process_root(root: Path) -> None:
    excel = root / "Excel"
    csv_dir = root / "Csv"
    diff_xlsx = excel / DIFF_XLSX
    write_difficulty_xlsx(diff_xlsx)
    bake_xlsx_to_csv(diff_xlsx, csv_dir / DIFF_CSV)

    op_xlsx = excel / OP_XLSX
    if not op_xlsx.is_file():
        raise SystemExit(f"missing {op_xlsx}")
    ensure_difficulty_id_column(op_xlsx)
    bake_xlsx_to_csv(op_xlsx, csv_dir / OP_CSV)


def main() -> None:
    process_root(ROOT)
    process_root(ROOT / "Mode2")
    print("Done.")


if __name__ == "__main__":
    main()
