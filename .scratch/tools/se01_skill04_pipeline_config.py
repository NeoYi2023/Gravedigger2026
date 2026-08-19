# -*- coding: utf-8 -*-
"""SE-01: add SkillEffectConfig columns + SkillEffect_04 rows + Skill_04 EffectImplemented=1; bake Mode2 CSV."""
from __future__ import annotations

import io
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables" / "Mode2"
EFFECT_XLSX = ROOT / "Excel" / "战斗_技能效果配置表_Combat_SkillEffectConfig.xlsx"
SKILL_XLSX = ROOT / "Excel" / "战斗_技能配置表_Combat_SkillConfig.xlsx"
EFFECT_CSV = ROOT / "Csv" / "Combat_SkillEffectConfig.csv"
SKILL_CSV = ROOT / "Csv" / "Combat_SkillConfig.csv"

SKILL04_EFFECTS = {
    "SkillEffect_04_1": ("OutgoingMulOnNewTargetFirstHit", "Mul=1.2", "OnOutgoingDamageSettle"),
    "SkillEffect_04_2": ("OutgoingMulOnNewTargetFirstHit", "Mul=1.3", "OnOutgoingDamageSettle"),
    "SkillEffect_04_3": ("OutgoingMulOnNewTargetFirstHit", "Mul=1.4", "OnOutgoingDamageSettle"),
    "SkillEffect_04_4": ("OutgoingMulOnNewTargetFirstHit", "Mul=1.5", "OnOutgoingDamageSettle"),
    "SkillEffect_04_5": ("OutgoingMulOnNewTargetFirstHit", "Mul=1.6", "OnOutgoingDamageSettle"),
}

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-10)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.10f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    return str(value)


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def bake(xlsx: Path, csv_out: Path) -> None:
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_csv_text(v) for v in row])
    wb.close()
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = next(i for i, row in enumerate(rows[:3]) if is_english_header(row))
    lines: list[str] = []
    for r_i, row in enumerate(rows[header_i:]):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"Baked {csv_out.name}: data_rows={len(lines) - 1}")


def patch_effect_config() -> None:
    wb = load_workbook(EFFECT_XLSX)
    ws = wb.active
    ws.cell(1, 3, "效果种类")
    ws.cell(2, 3, "登记制 PascalCase Token；空=未实现")
    ws.cell(3, 3, "EffectKind")
    ws.cell(1, 4, "效果参数")
    ws.cell(2, 4, "Key=Value|Key=Value|…")
    ws.cell(3, 4, "EffectParams")
    ws.cell(1, 5, "触发钩子")
    ws.cell(2, 5, "管线插入点枚举")
    ws.cell(3, 5, "TriggerHook")

    updated = 0
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None:
            continue
        sid = str(sid).strip()
        if sid not in SKILL04_EFFECTS:
            continue
        kind, params, hook = SKILL04_EFFECTS[sid]
        ws.cell(r, 3, kind)
        ws.cell(r, 4, params)
        ws.cell(r, 5, hook)
        updated += 1
        print(f"PATCH {sid} -> {kind} {params}")

    wb.save(EFFECT_XLSX)
    print(f"Saved {EFFECT_XLSX.name}: updated {updated} SkillEffect_04 rows")


def patch_skill_config() -> None:
    wb = load_workbook(SKILL_XLSX)
    ws = wb.active
    updated = 0
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if str(sid).strip() != "Skill_04":
            continue
        ws.cell(r, 12, 1)
        updated += 1
    wb.save(SKILL_XLSX)
    print(f"Saved {SKILL_XLSX.name}: Skill_04 EffectImplemented=1 for {updated} rows")


def main() -> None:
    patch_effect_config()
    patch_skill_config()
    bake(EFFECT_XLSX, EFFECT_CSV)
    bake(SKILL_XLSX, SKILL_CSV)


if __name__ == "__main__":
    main()
