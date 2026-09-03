# -*- coding: utf-8 -*-
"""Fix SearchExtract wave duplicate key (option B).

Point 1 × WaveIndex 1..4; FirstDelay=5; Interval=5; SP_01..04; Monster_01×10.
Excel source + Mode2 CSV bake (SPEC_04 §14.7).
"""
from __future__ import annotations

import io
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables" / "Mode2"
XLSX = ROOT / "Excel" / "搜打撤_刷怪波次配置表_SearchExtract_SearchExtractWaveSpawnConfig.xlsx"
CSV_PATH = ROOT / "Csv" / "SearchExtract_SearchExtractWaveSpawnConfig.csv"

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_MAX_DECIMAL_PLACES = 10

ROWS = [
    ["SearchExtract_01", 1, 1, 5, 5, "SP_01", "Monster_01", 10],
    ["SearchExtract_01", 1, 2, 5, 5, "SP_02", "Monster_01", 10],
    ["SearchExtract_01", 1, 3, 5, 5, "SP_03", "Monster_01", 10],
    ["SearchExtract_01", 1, 4, 5, 5, "SP_04", "Monster_01", 10],
]


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    return str(value)


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def write_xlsx() -> None:
    wb = load_workbook(XLSX)
    ws = wb.active
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    for row in ROWS:
        ws.append(row)
    wb.save(XLSX)


def bake_csv() -> None:
    with open(XLSX, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_csv_text(v) for v in row])
    wb.close()

    header_i = None
    for i, row in enumerate(rows[:3]):
        saw = False
        ok = True
        for cell in row:
            if cell is None or str(cell).strip() == "":
                continue
            saw = True
            if not EN_COL.match(str(cell).strip()):
                ok = False
                break
        if saw and ok:
            header_i = i
            break
    if header_i is None:
        raise RuntimeError("no English header in first 3 rows")

    export = rows[header_i:]
    max_cols = max((len(r) for r in export), default=0)
    for r in export:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))

    lines: list[str] = []
    for r_i, row in enumerate(export):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    text = "\n".join(lines) + ("\n" if lines else "")
    CSV_PATH.write_text(text, encoding="utf-8", newline="\n")


def main() -> None:
    write_xlsx()
    bake_csv()
    print(f"Wrote {len(ROWS)} wave rows → {XLSX.name} + {CSV_PATH.name}")
    print(CSV_PATH.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
