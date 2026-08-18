# -*- coding: utf-8 -*-
"""Mirror ConfigTableBaker Approach B: Excel → CSV (strip ≤2 doc rows).

Numeric emit follows SPEC_04 §14.6 (no binary float noise).
"""
from __future__ import annotations

import io
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
EXCEL_DIR = ROOT / "Excel"
CSV_DIR = ROOT / "Csv"

EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
MAX_SCAN = 3
_MAX_DECIMAL_PLACES = 10


def excel_to_csv_base(stem: str) -> str:
    parts = stem.split("_")
    if len(parts) != 4:
        raise ValueError(f"bad name: {stem}")
    return f"{parts[2]}_{parts[3]}"


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def find_header_index(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows[:MAX_SCAN]):
        if is_english_header(row):
            return i
    raise ValueError("no English header in first 3 rows")


def format_numeric_for_csv(number: float) -> str:
    """SPEC_04 §14.6: integer-valued → int string; else Round(10, AwayFromZero) + trim zeros."""
    if not math.isfinite(number):
        return str(number)

    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))

    # MidpointRounding.AwayFromZero ≈ Decimal ROUND_HALF_UP (ties away from zero)
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)

    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))

    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def sanitize_embedded_float_noise(text: str) -> str:
    """Rewrite float-noise literals inside encoded string fields (SPEC_04 §14.6)."""
    if not text or "." not in text:
        return text

    def repl(match: re.Match[str]) -> str:
        token = match.group(0)
        frac = token.split(".", 1)[1]
        looks_noisy = len(frac) > 10 or "000000" in frac or "999999" in frac
        if not looks_noisy:
            return token
        try:
            number = float(token)
        except ValueError:
            return token
        return format_numeric_for_csv(number)

    return re.sub(r"-?\d+\.\d+", repl, text)


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    if isinstance(value, Decimal):
        return format_numeric_for_csv(float(value))
    return sanitize_embedded_float_noise(str(value))


def read_sheet_rows(xlsx_path: Path) -> list[list[str]]:
    # Snapshot bytes first so Excel sharing the file does not abort (SPEC_04 §14.4).
    with open(xlsx_path, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [cell_to_csv_text(v) for v in row]
        rows.append(cells)
    wb.close()
    if not rows:
        return rows
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    return rows


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def build_csv(rows: list[list[str]]) -> str:
    lines: list[str] = []
    for r_i, row in enumerate(rows):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    return "\n".join(lines) + ("\n" if lines else "")


def main() -> None:
    changed = 0
    identical = 0
    for xlsx in sorted(EXCEL_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        csv_base = excel_to_csv_base(xlsx.stem)
        rows = read_sheet_rows(xlsx)
        header_i = find_header_index(rows)
        export = rows[header_i:]
        text = build_csv(export)
        out = CSV_DIR / f"{csv_base}.csv"
        old = out.read_text(encoding="utf-8") if out.is_file() else None
        out.write_text(text, encoding="utf-8", newline="\n")
        if old == text:
            identical += 1
            print(f"SAME {xlsx.name} → {out.name} (doc_rows={header_i})")
        else:
            changed += 1
            old_h = (old or "").splitlines()[:1]
            new_h = text.splitlines()[:1]
            print(f"DIFF {xlsx.name} → {out.name} (doc_rows={header_i}) header {old_h} -> {new_h}")
    print(f"Done. identical={identical} changed={changed}")


if __name__ == "__main__":
    main()
