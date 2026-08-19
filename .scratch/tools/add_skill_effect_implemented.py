# -*- coding: utf-8 -*-
"""Add EffectImplemented column to Combat_SkillConfig (Mode1+Mode2) and bake CSV."""
from __future__ import annotations

import io
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
IMPLEMENTED = {"Skill_01", "Skill_02", "Skill_03"}
EN_COL = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
COL = 12


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-10)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.10f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def sanitize_embedded_float_noise(text: str) -> str:
    if not text or "." not in text:
        return text

    def repl(match: re.Match[str]) -> str:
        token = match.group(0)
        frac = token.split(".", 1)[1]
        looks_noisy = len(frac) > 10 or "000000" in frac or "999999" in frac
        if not looks_noisy:
            return token
        try:
            number = float(token)
        except ValueError:
            return token
        return format_numeric_for_csv(number)

    return re.sub(r"-?\d+\.\d+", repl, text)


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    if isinstance(value, Decimal):
        return format_numeric_for_csv(float(value))
    return sanitize_embedded_float_noise(str(value))


def escape_csv_field(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n", "\r"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def is_english_header(row: list[str]) -> bool:
    saw = False
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        saw = True
        if not EN_COL.match(str(cell).strip()):
            return False
    return saw


def bake(xlsx: Path, csv_out: Path) -> None:
    with open(xlsx, "rb") as raw:
        data = raw.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_to_csv_text(v) for v in row])
    wb.close()
    max_cols = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))
    header_i = None
    for i, row in enumerate(rows[:3]):
        if is_english_header(row):
            header_i = i
            break
    if header_i is None:
        raise SystemExit(f"no English header: {xlsx}")
    lines: list[str] = []
    for r_i, row in enumerate(rows[header_i:]):
        if r_i > 0 and all(not (c or "").strip() for c in row):
            continue
        lines.append(",".join(escape_csv_field(c or "") for c in row))
    csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"Baked {csv_out.name}: data_rows={len(lines) - 1} last_col={rows[header_i][-1]}")


def patch(xlsx: Path, implemented_ids: set[str]) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active
    ws.cell(1, COL, "效果已实现")
    ws.cell(2, COL, "Demo 战斗效果是否已接线；不驱动战斗；UI-021 绿/红")
    ws.cell(3, COL, "EffectImplemented")
    n1 = n0 = 0
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None or str(sid).strip() == "":
            continue
        val = 1 if str(sid).strip() in implemented_ids else 0
        ws.cell(r, COL, val)
        if val:
            n1 += 1
        else:
            n0 += 1
    wb.save(xlsx)
    print(f"Saved {xlsx.name}: implemented={n1} unimplemented={n0}")


def main() -> None:
    m2 = ROOT / "Mode2" / "Excel" / "战斗_技能配置表_Combat_SkillConfig.xlsx"
    m1 = ROOT / "Excel" / "战斗_技能配置表_Combat_SkillConfig.xlsx"
    patch(m2, IMPLEMENTED)
    patch(m1, set())
    bake(m2, ROOT / "Mode2" / "Csv" / "Combat_SkillConfig.csv")
    bake(m1, ROOT / "Csv" / "Combat_SkillConfig.csv")


if __name__ == "__main__":
    main()
