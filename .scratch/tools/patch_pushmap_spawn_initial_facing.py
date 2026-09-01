# -*- coding: utf-8 -*-
"""Add InitialFacing column to PushMap spawn Excel + CSV (v0.83.58)."""
from pathlib import Path

import openpyxl

ROOT = Path(r"e:\Work\Cursor\Gravedigger2026\Gravedigger2026\Gravedigger2026\Assets\ConfigTables")
EXCELS = [
    ROOT / "Excel" / "推图战_刷怪配置表_PushMap_PushMapSpawnConfig.xlsx",
    ROOT / "Mode2" / "Excel" / "推图战_刷怪配置表_PushMap_PushMapSpawnConfig.xlsx",
]
CSVS = [
    ROOT / "Csv" / "PushMap_PushMapSpawnConfig.csv",
    ROOT / "Mode2" / "Csv" / "PushMap_PushMapSpawnConfig.csv",
]


def patch_xlsx(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Find English header row (row with GameplayConfigId)
    header_row = None
    for r in range(1, min(4, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v == "GameplayConfigId":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError(f"English header not found: {path}")

    # Check if already present
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        if ws.cell(header_row, c).value == "InitialFacing":
            print(f"skip (already has InitialFacing): {path}")
            return

    new_col = max_col + 1

    # Chinese name / description rows above English header (Approach B)
    if header_row >= 2:
        ws.cell(header_row - 1, new_col).value = (
            "0=每怪随机1~8；1正上 2右上 3正右 4右下 5正下 6左下 7正左 8左上；缺省5"
        )
    if header_row >= 3:
        # row1 = names, row2 = descriptions typically; if only 2 header rows before data:
        # Actually: row1 Chinese names, row2 Chinese desc, row3 English
        pass

    # Detect layout: if header_row==3, row1=names, row2=desc
    if header_row == 3:
        ws.cell(1, new_col).value = "初始朝向"
        ws.cell(2, new_col).value = (
            "0=每怪随机1~8；1正上/2右上/3正右/4右下/5正下/6左下/7正左/8左上；缺省5"
        )
    elif header_row == 2:
        ws.cell(1, new_col).value = "初始朝向"
    elif header_row == 1:
        pass

    ws.cell(header_row, new_col).value = "InitialFacing"

    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or str(first).strip() == "":
            continue
        ws.cell(r, new_col).value = 5

    wb.save(path)
    print(f"patched xlsx: {path}")


def patch_csv(path: Path) -> None:
    text = path.read_text(encoding="utf-8-sig")
    lines = text.splitlines()
    if not lines:
        raise RuntimeError(f"empty csv: {path}")
    header = lines[0]
    if "InitialFacing" in header.split(","):
        print(f"skip csv (already has InitialFacing): {path}")
        return
    out = [header + ",InitialFacing"]
    for line in lines[1:]:
        if not line.strip():
            continue
        out.append(line + ",5")
    path.write_text("\n".join(out) + "\n", encoding="utf-8")
    print(f"patched csv: {path}")


def dump_preview(path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    print("--- preview", path.name)
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        print(r, vals)
    wb.close()


def main() -> None:
    for p in EXCELS:
        if not p.exists():
            print("MISSING", p)
            continue
        dump_preview(p)
        patch_xlsx(p)
        dump_preview(p)
    for p in CSVS:
        if not p.exists():
            print("MISSING", p)
            continue
        patch_csv(p)


if __name__ == "__main__":
    main()
