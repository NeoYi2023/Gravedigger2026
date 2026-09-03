# -*- coding: utf-8 -*-
"""Restore SearchExtract wave sample to SPEC_04 §9.33 (v0.83.93).

Orders 1–2 × 2 waves; FirstDelay=2; Interval=8; Monster_01×3; SP_01/SP_02.
Excel source + Mode2 CSV bake (SPEC_04 §14.7).
"""
from __future__ import annotations

import csv
import math
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables" / "Mode2"
XLSX = ROOT / "Excel" / "搜打撤_刷怪波次配置表_SearchExtract_SearchExtractWaveSpawnConfig.xlsx"
CSV_PATH = ROOT / "Csv" / "SearchExtract_SearchExtractWaveSpawnConfig.csv"

_MAX_DECIMAL_PLACES = 10

WV_ZH = [
    "玩法配置ID",
    "搜集点序号",
    "波次序号",
    "第一波前置秒",
    "波间间隔秒",
    "刷怪点ID",
    "怪物ID",
    "出现数量",
]
WV_NOTE = [
    "FK → SearchExtractGameplayConfig",
    "对齐地图 ObjectiveOrder",
    "≥1；一行一波；同点升序",
    "自点激活起；同点各波须相同",
    "第 2 波起；同点各波须相同",
    "FK 地图 SpawnPoint",
    "FK → MonsterConfig",
    "≥1",
]
WV_EN = [
    "GameplayConfigId",
    "GatherPointOrder",
    "WaveIndex",
    "FirstWaveDelaySeconds",
    "WaveIntervalSeconds",
    "SpawnPointId",
    "MonsterId",
    "SpawnCount",
]
WV_ROWS = [
    ["SearchExtract_01", 1, 1, 2, 8, "SP_01", "Monster_01", 3],
    ["SearchExtract_01", 1, 2, 2, 8, "SP_01", "Monster_01", 3],
    ["SearchExtract_01", 2, 1, 2, 8, "SP_02", "Monster_01", 3],
    ["SearchExtract_01", 2, 2, 2, 8, "SP_02", "Monster_01", 3],
]


def format_numeric_for_csv(number: float) -> str:
    if not math.isfinite(number):
        return str(number)
    if abs(number - round(number)) < 1e-9 and abs(number) < 1e15:
        return str(int(round(number)))
    quant = Decimal(1).scaleb(-_MAX_DECIMAL_PLACES)
    rounded_dec = Decimal(repr(number)).quantize(quant, rounding=ROUND_HALF_UP)
    rounded = float(rounded_dec)
    if abs(rounded - round(rounded)) < 1e-12 and abs(rounded) < 1e15:
        return str(int(round(rounded)))
    text = f"{rounded:.{_MAX_DECIMAL_PLACES}f}".rstrip("0").rstrip(".")
    return text if text not in ("", "-") else "0"


def cell_to_csv_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return format_numeric_for_csv(value)
    return str(value)


def write_xlsx() -> None:
    if XLSX.exists():
        wb = load_workbook(XLSX)
        ws = wb.active
        # Keep three-row header if present; replace data rows only.
        # Always rewrite full sheet with canonical three-row header + sample.
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    ws.delete_rows(1, ws.max_row)
    ws.append(WV_ZH)
    ws.append(WV_NOTE)
    ws.append(WV_EN)
    for row in WV_ROWS:
        ws.append(row)
    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)


def write_csv() -> None:
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(WV_EN)
        for row in WV_ROWS:
            w.writerow([cell_to_csv_text(c) for c in row])


def main() -> None:
    write_xlsx()
    write_csv()
    print(f"Restored {len(WV_ROWS)} wave rows → {XLSX.name} + {CSV_PATH.name}")


if __name__ == "__main__":
    main()
