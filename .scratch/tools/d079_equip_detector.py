# -*- coding: utf-8 -*-
"""D-079: Equip_Detector EquipEffect + ItemCatalog + ShopPool (Mode1+Mode2)."""
from __future__ import annotations

import csv
import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
sys.path.insert(0, str(Path(__file__).resolve().parent))
from bake_config_tables_py import (  # noqa: E402
    build_csv,
    excel_to_csv_base,
    find_header_index,
    read_sheet_rows,
)

EQUIP_XLSX = "主角_装备配置表_Protagonist_ProtagonistEquipmentConfig.xlsx"
ITEM_XLSX = "通用_道具汇总表_Item_ItemCatalogConfig.xlsx"
SHOP_XLSX = "商店_商店商品池配置表_Shop_ShopPoolConfig.xlsx"

EFFECTS = {
    1: "DigProcessSpawnCountBonus_1",
    2: "DigProcessSpawnCountBonus_2",
    3: "DigProcessSpawnCountBonus_3",
    4: "DigProcessSpawnCountBonus_4",
    5: "DigProcessSpawnCountBonus_5",
}

DESC = "每级增加过程生成坟墓数量+1"

MODE1_ROWS = [
    ["Equip_Detector", "1", "探测器", "Icon_Detector", "1", "1", "Dig", EFFECTS[1], f"{DESC}（1级）"],
    ["Equip_Detector", "2", "探测器", "Icon_Detector", "1", "1", "Dig", EFFECTS[2], f"{DESC}（2级）"],
    ["Equip_Detector", "3", "探测器", "Icon_Detector", "1", "1", "Dig", EFFECTS[3], f"{DESC}（3级）"],
    ["Equip_Detector", "4", "探测器", "Icon_Detector", "1", "1", "Dig", EFFECTS[4], f"{DESC}（4级）"],
    ["Equip_Detector", "5", "探测器", "Icon_Detector", "", "1", "Dig", EFFECTS[5], f"{DESC}（满级）"],
]

MODE2_CONVERT = {1: "1", 2: "2", 3: "3", 4: "4", 5: "5"}

ITEM_ROW = [
    "Equip_Detector",
    "探测器",
    "Icon_Detector",
    "ProtagonistEquipment",
    "Protagonist_ProtagonistEquipmentConfig",
    "增加过程生成坟墓数量的主角装备",
    "20",
]


def find_en_header(ws) -> tuple[int, dict[str, int]]:
    for r in range(1, min(4, ws.max_row + 1)):
        cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        texts = [str(v).strip() if v is not None else "" for v in cells]
        if any(t in ("EquipId", "ItemId", "ShopPoolId") for t in texts):
            cols = {t: i + 1 for i, t in enumerate(texts) if t}
            return r, cols
    raise SystemExit("no EN header (EquipId/ItemId/ShopPoolId) in first 3 rows")


def patch_equip_xlsx(path: Path, mode: str) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_row, cols = find_en_header(ws)
    id_col = cols["EquipId"]
    lvl_col = cols["EquipLevel"]
    effect_col = cols["EquipEffect"]
    desc_col = cols.get("Description")

    existing_levels: dict[int, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        eid = ws.cell(r, id_col).value
        if eid != "Equip_Detector":
            continue
        lvl_raw = ws.cell(r, lvl_col).value
        if lvl_raw is None or str(lvl_raw).strip() == "":
            continue
        lvl = int(lvl_raw)
        existing_levels[lvl] = r

    if existing_levels:
        for lvl, row_i in existing_levels.items():
            ws.cell(row_i, effect_col).value = EFFECTS[lvl]
            if desc_col:
                ws.cell(row_i, desc_col).value = f"{DESC}（{lvl}级）" if lvl < 5 else f"{DESC}（满级）"
        print(f"PATCHED EquipEffect on {len(existing_levels)} rows: {path}")
    else:
        rows = MODE1_ROWS if mode == "Mode1" else [
            [
                "Equip_Detector",
                str(lvl),
                "探测器",
                "Icon_Detector",
                "" if lvl == 5 else "1",
                MODE2_CONVERT[lvl],
                "Dig",
                EFFECTS[lvl],
                f"{DESC}（{lvl}级）" if lvl < 5 else f"{DESC}（满级）",
            ]
            for lvl in range(1, 6)
        ]
        for row in rows:
            ws.append(row)
        print(f"APPENDED 5 Equip_Detector rows: {path}")

    wb.save(path)


def bake_one(excel_dir: Path, csv_dir: Path, xlsx_name: str) -> None:
    xlsx = excel_dir / xlsx_name
    if not xlsx.is_file():
        print(f"SKIP bake missing {xlsx}")
        return
    rows = read_sheet_rows(xlsx)
    header_i = find_header_index(rows)
    export = rows[header_i:]
    text = build_csv(export)
    out = csv_dir / f"{excel_to_csv_base(xlsx.stem)}.csv"
    out.write_text(text, encoding="utf-8", newline="\n")
    print(f"BAKE {xlsx.name} → {out.name}")


def ensure_item_csv(csv_path: Path) -> None:
    text = csv_path.read_text(encoding="utf-8")
    if "Equip_Detector" in text:
        print(f"SKIP ItemCatalog already has Detector: {csv_path}")
        return
    lines = text.rstrip("\n").splitlines()
    lines.append(",".join(ITEM_ROW))
    csv_path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"APPEND ItemCatalog: {csv_path}")


def ensure_item_xlsx(xlsx_path: Path) -> None:
    if not xlsx_path.is_file():
        print(f"SKIP missing Item xlsx: {xlsx_path}")
        return
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header_row, cols = find_en_header(ws)
    id_col = cols["ItemId"]
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, id_col).value == "Equip_Detector":
            print(f"SKIP Item xlsx already has Detector: {xlsx_path}")
            return
    ws.append(ITEM_ROW)
    wb.save(xlsx_path)
    print(f"APPEND Item xlsx: {xlsx_path}")


def patch_shop_pool(csv_path: Path, xlsx_path: Path | None) -> None:
    rows: list[list[str]] = []
    with csv_path.open(encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        raise SystemExit(f"empty shop csv {csv_path}")
    header = rows[0]
    items_i = header.index("PoolItemsRaw")
    changed = False
    for row in rows[1:]:
        if len(row) <= items_i:
            continue
        raw = row[items_i]
        if "Equip_Detector" in raw:
            continue
        # Prefixed into Pool_01 (and others for visibility): after Equip_IronShovel or at end
        if "Equip_IronShovel;A;10" in raw:
            row[items_i] = raw.replace(
                "Equip_IronShovel;A;10",
                "Equip_IronShovel;A;10|Equip_Detector;A;10",
                1,
            )
        else:
            row[items_i] = raw + "|Equip_Detector;A;10"
        changed = True
    if changed:
        with csv_path.open("w", encoding="utf-8", newline="\n") as f:
            writer = csv.writer(f, lineterminator="\n")
            writer.writerows(rows)
        print(f"PATCH ShopPool csv: {csv_path}")
    else:
        print(f"SKIP ShopPool csv already has Detector: {csv_path}")

    if xlsx_path is None or not xlsx_path.is_file():
        print(f"SKIP ShopPool xlsx missing: {xlsx_path}")
        return
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header_row, cols = find_en_header(ws)
    items_col = cols["PoolItemsRaw"]
    x_changed = False
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, items_col).value
        if raw is None:
            continue
        raw_s = str(raw)
        if "Equip_Detector" in raw_s:
            continue
        if "Equip_IronShovel;A;10" in raw_s:
            ws.cell(r, items_col).value = raw_s.replace(
                "Equip_IronShovel;A;10",
                "Equip_IronShovel;A;10|Equip_Detector;A;10",
                1,
            )
        else:
            ws.cell(r, items_col).value = raw_s + "|Equip_Detector;A;10"
        x_changed = True
    if x_changed:
        wb.save(xlsx_path)
        print(f"PATCH ShopPool xlsx: {xlsx_path}")
    else:
        print(f"SKIP ShopPool xlsx already has Detector: {xlsx_path}")


def main() -> None:
    mode1_equip = ROOT / "Excel" / EQUIP_XLSX
    mode2_equip = ROOT / "Mode2" / "Excel" / EQUIP_XLSX
    patch_equip_xlsx(mode1_equip, "Mode1")
    patch_equip_xlsx(mode2_equip, "Mode2")
    bake_one(ROOT / "Excel", ROOT / "Csv", EQUIP_XLSX)
    bake_one(ROOT / "Mode2" / "Excel", ROOT / "Mode2" / "Csv", EQUIP_XLSX)

    ensure_item_csv(ROOT / "Csv" / "Item_ItemCatalogConfig.csv")
    ensure_item_csv(ROOT / "Mode2" / "Csv" / "Item_ItemCatalogConfig.csv")
    ensure_item_xlsx(ROOT / "Excel" / ITEM_XLSX)
    ensure_item_xlsx(ROOT / "Mode2" / "Excel" / ITEM_XLSX)
    # Re-bake item if xlsx exists so CSV matches Excel
    bake_one(ROOT / "Excel", ROOT / "Csv", ITEM_XLSX)
    bake_one(ROOT / "Mode2" / "Excel", ROOT / "Mode2" / "Csv", ITEM_XLSX)

    patch_shop_pool(
        ROOT / "Mode2" / "Csv" / "Shop_ShopPoolConfig.csv",
        ROOT / "Mode2" / "Excel" / SHOP_XLSX,
    )


if __name__ == "__main__":
    main()
