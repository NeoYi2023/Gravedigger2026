# -*- coding: utf-8 -*-
"""SE-03: Mode2 Level_01 last-combat fork Opt_SE_Demo_01 + rewrite operation table to OptionId1..5.

Excel three-row headers (SPEC_04 §14.7). Also writes script .meta for new SearchExtract files.
"""
from __future__ import annotations

import csv
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
CFG = ROOT / "Gravedigger2026" / "Assets" / "ConfigTables" / "Mode2"
SCRIPTS = ROOT / "Gravedigger2026" / "Assets" / "Scripts"

OP_XLSX = "关卡_关卡运作表_Level_LevelOperationConfig.xlsx"
OP_CSV = "Level_LevelOperationConfig.csv"
SUB_XLSX = "关卡_子关卡表_Level_SubLevelConfig.xlsx"
SUB_CSV = "Level_SubLevelConfig.csv"

OP_ZH = ["关卡ID", "阶段编号", "玩法选项ID1", "玩法选项ID2", "玩法选项ID3", "玩法选项ID4", "玩法选项ID5"]
OP_NOTE = [
    "同 ID 多行 = 该关全部 Stage",
    "同关升序；一行 = 一个 Stage",
    "FK → SubLevelConfig；空=无",
    "同 Stage 最多 5 套",
    "",
    "",
    "",
]
OP_EN = [
    "LevelId",
    "StageNumber",
    "GameplayOptionId1",
    "GameplayOptionId2",
    "GameplayOptionId3",
    "GameplayOptionId4",
    "GameplayOptionId5",
]

META_SCRIPT = """fileFormatVersion: 2
guid: {guid}
MonoImporter:
  externalObjects: {{}}
  serializedVersion: 2
  defaultReferences: []
  executionOrder: 0
  icon: {{instanceID: 0}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""

META_FOLDER = """fileFormatVersion: 2
guid: {guid}
folderAsset: yes
DefaultImporter:
  externalObjects: {{}}
  userData: 
  assetBundleName: 
  assetBundleVariant: 
"""


def ensure_meta(path: Path, folder: bool = False) -> None:
    meta = path.with_suffix(path.suffix + ".meta") if path.suffix else Path(str(path) + ".meta")
    if path.is_dir():
        meta = Path(str(path) + ".meta")
    if meta.exists():
        return
    template = META_FOLDER if folder or path.is_dir() else META_SCRIPT
    meta.write_text(template.format(guid=uuid.uuid4().hex), encoding="utf-8")
    print(f"wrote {meta.name}")


def write_op_xlsx(path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OP_ZH)
    ws.append(OP_NOTE)
    ws.append(OP_EN)
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def patch_sub_unlock(xlsx_path: Path, csv_path: Path) -> None:
    wanted = "Opt_L01_S5_PushMap|Opt_SE_Demo_01"
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
        try:
            id_col = headers.index("GameplayOptionId") + 1
            unlock_col = headers.index("UnlockNextOptionIds") + 1
        except ValueError as exc:
            raise SystemExit(f"SubLevel excel missing column: {exc}") from exc
        patched = False
        for row in ws.iter_rows(min_row=4):
            oid = row[id_col - 1].value
            if oid == "Opt_L01_S4_UpgradeManufacture":
                ws.cell(row[0].row, unlock_col, wanted)
                patched = True
                break
        if not patched:
            raise SystemExit("Opt_L01_S4_UpgradeManufacture not found in SubLevel excel")
        wb.save(xlsx_path)
        print(f"patched excel unlock → {wanted}")
    else:
        print(f"skip missing {xlsx_path}")

    # CSV already edited by agent; verify.
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        for raw in csv.DictReader(f):
            if raw.get("GameplayOptionId") == "Opt_L01_S4_UpgradeManufacture":
                got = (raw.get("UnlockNextOptionIds") or "").strip()
                if got != wanted:
                    raise SystemExit(f"SubLevel CSV unlock mismatch: {got!r}")
                print("csv unlock ok")
                return
    raise SystemExit("UM row missing in SubLevel CSV")


def main() -> None:
    csv_dir = CFG / "Csv"
    excel = CFG / "Excel"
    op_csv = csv_dir / OP_CSV
    rows: list[list[str]] = []
    with op_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            rows.append([raw.get(k) or "" for k in OP_EN])
    try:
        write_op_xlsx(excel / OP_XLSX, rows)
        print(f"wrote {OP_XLSX} rows={len(rows)}")
    except PermissionError:
        print(f"EXCEL LOCKED: close {OP_XLSX} and re-run; CSV already updated")

    try:
        patch_sub_unlock(excel / SUB_XLSX, csv_dir / SUB_CSV)
    except PermissionError:
        print(f"EXCEL LOCKED: close {SUB_XLSX} and re-run; CSV already updated")

    folders = [
        SCRIPTS / "Core" / "SearchExtract",
        SCRIPTS / "Gameplay" / "SearchExtract",
    ]
    files = [
        SCRIPTS / "Core" / "SearchExtract" / "SearchExtractPhase.cs",
        SCRIPTS / "Core" / "SearchExtract" / "SearchExtractSessionService.cs",
        SCRIPTS / "Core" / "Level" / "SearchExtractStageModule.cs",
        SCRIPTS / "Gameplay" / "SearchExtract" / "SearchExtractStageController.cs",
    ]
    for folder in folders:
        ensure_meta(folder, folder=True)
    for cs in files:
        ensure_meta(cs, folder=False)


if __name__ == "__main__":
    main()
