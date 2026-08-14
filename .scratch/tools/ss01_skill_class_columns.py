# -*- coding: utf-8 -*-
"""SS-01: SkillConfig IconAssetId + Skill_01 Lv2; ClassConfig DefaultSkillIds; then Bake."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2] / "Gravedigger2026" / "Assets" / "ConfigTables"
SKILL_XLSX = "战斗_技能配置表_Combat_SkillConfig.xlsx"
CLASS_XLSX = "制造_职业配置表_Manufacture_ClassConfig.xlsx"

# Reuse bake helpers (SPEC_04 §14.6).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from bake_config_tables_py import (  # noqa: E402
    build_csv,
    excel_to_csv_base,
    find_header_index,
    read_sheet_rows,
)


def find_en_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(c == "SkillId" or c == "ClassId" for c in cells):
            return i
    raise RuntimeError("no EN header in first 3 rows")


def header_list(ws, header_row: int) -> list[str]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    names = [str(c).strip() if c is not None else "" for c in row]
    while names and names[-1] == "":
        names.pop()
    return names


def insert_column_after(ws, header_row: int, after_en: str, en: str, zh: str, note: str) -> None:
    names = header_list(ws, header_row)
    if en in names:
        return
    if after_en not in names:
        raise RuntimeError(f"missing column {after_en} in {names}")
    insert_at = names.index(after_en) + 2  # 1-based, after after_en
    ws.insert_cols(insert_at)
    if header_row >= 3:
        ws.cell(1, insert_at, zh)
        ws.cell(2, insert_at, note)
    ws.cell(header_row, insert_at, en)


def append_column(ws, header_row: int, en: str, zh: str, note: str) -> None:
    names = header_list(ws, header_row)
    if en in names:
        return
    insert_at = len(names) + 1
    ws.insert_cols(insert_at)
    if header_row >= 3:
        ws.cell(1, insert_at, zh)
        ws.cell(2, insert_at, note)
    ws.cell(header_row, insert_at, en)


def set_cell_by_en(ws, header_row: int, data_row: int, en: str, value) -> None:
    names = header_list(ws, header_row)
    col = names.index(en) + 1
    ws.cell(data_row, col, value)


def find_data_row(ws, header_row: int, id_en: str, id_value: str, extra: tuple[str, object] | None = None) -> int | None:
    names = header_list(ws, header_row)
    id_col = names.index(id_en) + 1
    extra_col = names.index(extra[0]) + 1 if extra else None
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(r, id_col).value
        if cell is None or str(cell).strip() != id_value:
            continue
        if extra is None:
            return r
        extra_val = ws.cell(r, extra_col).value
        if extra_val is not None and str(extra_val).strip() == str(extra[1]):
            return r
    return None


def copy_row_values(ws, src: int, dst: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        ws.cell(dst, c, ws.cell(src, c).value)


def patch_skill(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_row = find_en_header_row(ws)
    insert_column_after(
        ws,
        header_row,
        "Description",
        "IconAssetId",
        "技能图标",
        "UI 图标资源 Id；缺/空=无图",
    )
    names = header_list(ws, header_row)
    lv2 = find_data_row(ws, header_row, "SkillId", "Skill_01", ("SkillLevel", 2))
    if lv2 is None:
        lv1 = find_data_row(ws, header_row, "SkillId", "Skill_01", ("SkillLevel", 1))
        if lv1 is None:
            raise RuntimeError(f"{path}: missing Skill_01 Lv1")
        insert_at = lv1 + 1
        ws.insert_rows(insert_at)
        copy_row_values(ws, lv1, insert_at, len(names))
        set_cell_by_en(ws, header_row, insert_at, "SkillLevel", 2)
        set_cell_by_en(ws, header_row, insert_at, "IconAssetId", "")
        set_cell_by_en(ws, header_row, insert_at, "LossOfControlChanceBonus", 0.03)
    wb.save(path)
    print(f"OK skill {path}")


def patch_class(path: Path, warrior_ids: set[str]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_row = find_en_header_row(ws)
    append_column(
        ws,
        header_row,
        "DefaultSkillIds",
        "制造默认获得技能ID",
        "空=无；SkillId 或 SkillId|SkillId；FK → SkillConfig.SkillId",
    )
    names = header_list(ws, header_row)
    id_col = names.index("ClassId") + 1
    skill_col = names.index("DefaultSkillIds") + 1
    for r in range(header_row + 1, ws.max_row + 1):
        class_id = ws.cell(r, id_col).value
        if class_id is None:
            continue
        cid = str(class_id).strip()
        if not cid:
            continue
        ws.cell(r, skill_col, "Skill_01" if cid in warrior_ids else "")
    wb.save(path)
    print(f"OK class {path}")


def bake_one(excel_dir: Path, csv_dir: Path, xlsx_name: str) -> None:
    xlsx = excel_dir / xlsx_name
    rows = read_sheet_rows(xlsx)
    header_i = find_header_index(rows)
    text = build_csv(rows[header_i:])
    out = csv_dir / f"{excel_to_csv_base(xlsx.stem)}.csv"
    out.write_text(text, encoding="utf-8", newline="\n")
    print(f"BAKE {xlsx.name} → {out.name}")


def main() -> None:
    mode1_excel = ROOT / "Excel"
    mode1_csv = ROOT / "Csv"
    mode2_excel = ROOT / "Mode2" / "Excel"
    mode2_csv = ROOT / "Mode2" / "Csv"

    patch_skill(mode1_excel / SKILL_XLSX)
    patch_skill(mode2_excel / SKILL_XLSX)
    patch_class(mode1_excel / CLASS_XLSX, {"Class_Warrior"})
    patch_class(mode2_excel / CLASS_XLSX, {"Class_Warrior", "Class_Warrior_0"})

    for excel_dir, csv_dir in ((mode1_excel, mode1_csv), (mode2_excel, mode2_csv)):
        bake_one(excel_dir, csv_dir, SKILL_XLSX)
        bake_one(excel_dir, csv_dir, CLASS_XLSX)


if __name__ == "__main__":
    main()
