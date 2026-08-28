# -*- coding: utf-8 -*-
"""Rewrite Defend_MonsterConfig Excel rows 1–2 to match row-3 English columns.

Does not touch data rows or CSV. Run: py -3.11 fix_monster_config_excel_headers.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
from migrate_config_excel_3row_headers import FIELDS, zh_for  # noqa: E402

CSV_BASE = "Defend_MonsterConfig"
XLSX_PATHS = [
    ROOT.parents[1]
    / "Gravedigger2026"
    / "Assets"
    / "ConfigTables"
    / "Excel"
    / "防守_怪物配置表_Defend_MonsterConfig.xlsx",
    ROOT.parents[1]
    / "Gravedigger2026"
    / "Assets"
    / "ConfigTables"
    / "Mode2"
    / "Excel"
    / "防守_怪物配置表_Defend_MonsterConfig.xlsx",
]


def patch(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    max_col = ws.max_column
    en_names: list[str] = []
    for col in range(1, max_col + 1):
        value = ws.cell(3, col).value
        en = "" if value is None else str(value).strip()
        en_names.append(en)

    while en_names and not en_names[-1]:
        en_names.pop()

    missing = [en for en in en_names if en and en not in FIELDS[CSV_BASE]]
    if missing:
        raise SystemExit(f"{path.name}: columns not in FIELDS: {missing}")

    for col, en in enumerate(en_names, start=1):
        zh, note = zh_for(CSV_BASE, en)
        ws.cell(1, col).value = zh
        ws.cell(2, col).value = note

    wb.save(path)
    print(f"OK {path} ({len(en_names)} cols)")


def main() -> None:
    for path in XLSX_PATHS:
        if not path.is_file():
            raise SystemExit(f"missing: {path}")
        try:
            patch(path)
        except PermissionError:
            raise SystemExit(
                f"locked (close Excel first): {path}\n"
                f"lock file may exist: {path.parent / ('~$' + path.name)}"
            )


if __name__ == "__main__":
    main()
