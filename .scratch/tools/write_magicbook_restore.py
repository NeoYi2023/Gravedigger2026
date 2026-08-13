# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import load_workbook

bases = [
    Path(__file__).resolve().parents[2]
    / "Gravedigger2026"
    / "Assets"
    / "ConfigTables",
    Path(__file__).resolve().parents[2]
    / "Gravedigger2026"
    / "Assets"
    / "ConfigTables"
    / "Mode2",
]
row = [
    "MagicBook_Restore",
    1,
    "SoldierManufacture",
    "RaceWeightPick",
    "",
    "",
    "还原",
    "装备后制造时按部位 RaceId 加权随机定种族",
]
xlsx_name = "制造_魔法书配置表_Manufacture_MagicBookConfig.xlsx"
csv_name = "Manufacture_MagicBookConfig.csv"
csv_text = (
    "MagicBookId,IsUnique,EffectPhase,EffectPayload,EffectParams,IconAssetId,DisplayName,Description\n"
    "MagicBook_Restore,1,SoldierManufacture,RaceWeightPick,,,还原,装备后制造时按部位 RaceId 加权随机定种族\n"
)

for base in bases:
    xlsx = base / "Excel" / xlsx_name
    wb = load_workbook(xlsx)
    ws = wb.active
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    ws.append(row)
    wb.save(xlsx)
    print("wrote", xlsx)
    csv_path = base / "Csv" / csv_name
    csv_path.write_text(csv_text, encoding="utf-8")
    print("wrote", csv_path)
