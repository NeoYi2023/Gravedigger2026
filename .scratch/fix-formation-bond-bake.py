# -*- coding: utf-8 -*-
"""Fix FormationBond Excel three-row header and rebake CSV."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / ".scratch" / "tools"))
from bake_config_tables_py import build_csv, find_header_index, read_sheet_rows

EXCEL_PATH = (
    ROOT
    / "Gravedigger2026"
    / "Assets"
    / "ConfigTables"
    / "Excel"
    / "战斗_阵容羁绊配置表_Combat_FormationBondConfig.xlsx"
)
CSV_PATH = (
    ROOT / "Gravedigger2026" / "Assets" / "ConfigTables" / "Csv" / "Combat_FormationBondConfig.csv"
)

ROW1 = ["羁绊ID", "羁绊等级", "羁绊名称", "羁绊图标", "羁绊介绍", "羁绊激活条件", "羁绊Buff"]
ROW2 = [
    "复合主键之一",
    "复合主键之一；≥1；同 Id 多等级互斥（规则见 §3.17）",
    "UI 展示",
    "运行时 Resources/UI/Bonds/{IconAssetId}；缺图占位",
    "UI 只读；不作效果器",
    "结构化 DSL（§3.17）；加载期校验",
    "FK → SkillEffectConfig.SkillEffectId",
]


def main() -> None:
    with open(EXCEL_PATH, "rb") as f:
        data = f.read()

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    for col, val in enumerate(ROW1, 1):
        ws.cell(1, col, val)
    for col, val in enumerate(ROW2, 1):
        ws.cell(2, col, val)
    wb.save(EXCEL_PATH)
    wb.close()

    rows = read_sheet_rows(EXCEL_PATH)
    header_i = find_header_index(rows)
    csv_text = build_csv(rows[header_i:])
    CSV_PATH.write_text(csv_text, encoding="utf-8", newline="\n")
    print(f"header_index={header_i}")
    print(csv_text)


if __name__ == "__main__":
    main()
