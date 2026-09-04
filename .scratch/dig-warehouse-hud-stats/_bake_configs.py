# -*- coding: utf-8 -*-
import csv
import os
from openpyxl import Workbook, load_workbook

ROOT = r"e:\Work\Cursor\Gravedigger2026\Gravedigger2026\Gravedigger2026\Assets\ConfigTables"
MODE2_CSV = os.path.join(ROOT, "Mode2", "Csv")
MODE2_XLS = os.path.join(ROOT, "Mode2", "Excel")
MODE1_CSV = os.path.join(ROOT, "Csv")
MODE1_XLS = os.path.join(ROOT, "Excel")

CLASS_TO_BASE = {
    "Class_BaseWarrior": "战士",
    "Class_Warrior": "战士",
    "Class_Guardian": "战士",
    "Class_Berserker": "战士",
    "Class_Paladin": "战士",
    "Class_BaseArcher": "射手",
    "Class_Archer": "射手",
    "Class_BombMaster": "射手",
    "Class_Longbowman": "射手",
    "Class_BaseMage": "法师",
    "Class_Mage": "法师",
    "Class_IceMage": "法师",
    "Class_FireMage": "法师",
    "Class_DarkMage": "法师",
    "Class_BaseRogue": "刺客",
    "Class_Rogue": "刺客",
    "Class_Brawler": "刺客",
    "Class_Shadowblade": "刺客",
}
DESC_TO_BASE = {
    "战士": "战士",
    "射手": "射手",
    "法师": "法师",
    "刺客": "刺客",
    "近卫": "战士",
    "狂战士": "战士",
    "圣骑士": "战士",
    "炸弹师": "射手",
    "长弓手": "射手",
    "冰法": "法师",
    "火法": "法师",
    "暗黑法师": "法师",
    "格斗师": "刺客",
    "影刃": "刺客",
}


def base_for_row(row):
    is_ph = (row.get("IsPrimaryHand") or "0").strip()
    if is_ph != "1":
        return ""
    cr = (row.get("ClassRestrict") or "").strip().split("|")[0]
    if cr in CLASS_TO_BASE:
        return CLASS_TO_BASE[cr]
    desc = (row.get("Description") or "").strip()
    if desc in DESC_TO_BASE:
        return DESC_TO_BASE[desc]
    return ""


def main():
    bp_path = os.path.join(MODE2_CSV, "Manufacture_BodyPartConfig.csv")
    with open(bp_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    if "BaseClass" not in fieldnames:
        idx = fieldnames.index("ClassRestrict") + 1 if "ClassRestrict" in fieldnames else len(fieldnames)
        fieldnames.insert(idx, "BaseClass")

    for r in rows:
        r["BaseClass"] = base_for_row(r)

    with open(bp_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})

    print(
        "Mode2 BodyPart CSV updated, primary hands with BaseClass:",
        sum(1 for r in rows if r.get("IsPrimaryHand") == "1" and r.get("BaseClass")),
    )

    bp_xlsx = os.path.join(MODE2_XLS, "制造_躯体材料配置表_Manufacture_BodyPartConfig.xlsx")
    wb = load_workbook(bp_xlsx)
    ws = wb.active
    en_headers = [c.value for c in ws[3]]
    print("Excel EN headers:", en_headers)
    if "BaseClass" not in en_headers:
        if "ClassRestrict" in en_headers:
            col = en_headers.index("ClassRestrict") + 2
        else:
            col = len(en_headers) + 1
        ws.insert_cols(col)
        ws.cell(1, col, "基础职业")
        ws.cell(2, col, "主要手职业统计归桶；战士|射手|法师|刺客；非主要手可空")
        ws.cell(3, col, "BaseClass")
        en_headers = [c.value for c in ws[3]]

    bc_col = en_headers.index("BaseClass") + 1
    id_col = en_headers.index("BodyPartId") + 1
    by_id = {r["BodyPartId"]: r.get("BaseClass", "") for r in rows}
    for r_i in range(4, ws.max_row + 1):
        bid = ws.cell(r_i, id_col).value
        if not bid:
            continue
        ws.cell(r_i, bc_col, by_id.get(str(bid), ""))

    wb.save(bp_xlsx)
    print("Mode2 BodyPart Excel saved")

    loc_header = ["TextKey", "TextZh", "TextEn", "Comment"]
    loc_row = [
        "DigWarehouseHoverTips",
        "此处统计可大致制造的士兵种族与职业数量",
        "",
        "Dig HUD Warehouse hover tips",
    ]

    def write_loc_csv(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(loc_header)
            w.writerow(loc_row)

    def write_loc_xlsx(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet1"
        zh = ["文案键", "中文文案", "英文文案", "备注"]
        zh_desc = ["主键；代码引用", "Demo 权威展示串", "预留多语言", "策划说明不参与逻辑"]
        for i, (a, b, c) in enumerate(zip(zh, zh_desc, loc_header), 1):
            ws2.cell(1, i, a)
            ws2.cell(2, i, b)
            ws2.cell(3, i, c)
        for i, v in enumerate(loc_row, 1):
            ws2.cell(4, i, v)
        wb2.save(path)

    write_loc_csv(os.path.join(MODE2_CSV, "Common_LocalizedDescriptionConfig.csv"))
    write_loc_csv(os.path.join(MODE1_CSV, "Common_LocalizedDescriptionConfig.csv"))
    write_loc_xlsx(os.path.join(MODE2_XLS, "通用_多语言描述表_Common_LocalizedDescriptionConfig.xlsx"))
    os.makedirs(MODE1_XLS, exist_ok=True)
    write_loc_xlsx(os.path.join(MODE1_XLS, "通用_多语言描述表_Common_LocalizedDescriptionConfig.xlsx"))
    print("LocalizedDescription Excel+CSV written")

    missing = [r["BodyPartId"] for r in rows if r.get("IsPrimaryHand") == "1" and not r.get("BaseClass")]
    print("Missing BaseClass primary:", missing)


if __name__ == "__main__":
    main()
